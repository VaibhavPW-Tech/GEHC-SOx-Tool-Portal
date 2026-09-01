"""
Tech Assisted Audit Tools — Single-File Streamlit Application
=======================================================
Contains: SSO login, self-service account registration with admin
approval workflow, password change, Admin Portal (Admin-only),
Application Dashboard, Saviynt Tool, CM Automation Tool.

Demo credentials (auto-seeded on first run, ./data/users.json):
    Admin              -> admin1   / Admin@123
"""

import json
import os
import hashlib
import datetime
import re
import time
import glob
import zipfile
import tempfile
import shutil
import subprocess
import io

import streamlit as st
import pandas as pd
import pdfplumber

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

def now_ist() -> datetime.datetime:
    return datetime.datetime.now(IST)


# ============================================================================
# SECTION 1 — AUTH: user database, authentication, registration, approval
# ============================================================================
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
USERS_FILE = os.path.join(DATA_DIR, "users.json")

ROLE_ADMIN = "Admin"
ROLE_SOX   = "GEHC IT SOX Team"
ROLE_USER  = "Regular User"
ALL_ROLES = [ROLE_ADMIN, ROLE_SOX]

APP_SAVIYNT       = "Access Reconciliation Suite"
APP_CM_AUTOMATION = "CM Automation"
ALL_APPS = [APP_SAVIYNT, APP_CM_AUTOMATION]
SOX_PROVISIONABLE_APPS = [APP_SAVIYNT, APP_CM_AUTOMATION]

STATUS_PENDING  = "pending"
STATUS_APPROVED = "approved"
STATUS_REJECTED = "rejected"


def _hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _seed_users():
    return {
        "admin1": {"name": "Portal Administrator", "email": "admin1@gehealthcare.com",
                   "password_hash": _hash_password("Admin@123"), "role": ROLE_ADMIN,
                   "apps": ALL_APPS.copy(), "active": True, "status": STATUS_APPROVED,
                   "created_on": now_ist().isoformat()},
    }


def _ensure_data_file():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(_seed_users(), f, indent=2)


def load_users() -> dict:
    _ensure_data_file()
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        users = json.load(f)
    needs_resave = False
    for u in users.values():
        cleaned = [a for a in u.get("apps", []) if a in ALL_APPS]
        if cleaned != u.get("apps", []):
            u["apps"] = cleaned
            needs_resave = True
        if "status" not in u:
            u["status"] = STATUS_APPROVED
            needs_resave = True
    if needs_resave:
        save_users(users)
    return users


def save_users(users: dict):
    _ensure_data_file()
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2)


def get_user(sso_id: str):
    return load_users().get(sso_id.strip().lower())


def get_pending_users() -> dict:
    return {sid: u for sid, u in load_users().items() if u.get("status") == STATUS_PENDING}


def authenticate(sso_id: str, password: str):
    if not sso_id or not sso_id.strip():
        return False, "Please enter your SSO ID."
    if not password:
        return False, "Please enter your password."

    sso_key = sso_id.strip().lower()
    users = load_users()
    user = users.get(sso_key)

    if user is None:
        return False, "Invalid SSO ID or Password. Please re-enter correct details."
    if not user.get("active", True):
        return False, "This account has been deactivated. Please contact your Admin."
    if user.get("password_hash") != _hash_password(password):
        return False, "Invalid SSO ID or Password. Please re-enter correct details."
    if user.get("status") == STATUS_REJECTED:
        return False, "Your account request was rejected by the Admin. Please contact your Admin team for more details."

    user_out = dict(user)
    user_out["sso_id"] = sso_key
    return True, user_out


def register_user(sso_id: str, name: str, email: str, password: str):
    if not sso_id or not sso_id.strip():
        return False, "Please enter an SSO ID."
    if not name or not name.strip():
        return False, "Please enter your full name."
    if not email or not email.strip():
        return False, "Please enter your email."
    if not password or len(password) < 6:
        return False, "Password must be at least 6 characters."

    sso_key = sso_id.strip().lower()
    users = load_users()
    if sso_key in users:
        return False, f"SSO ID '{sso_key}' is already registered. Please sign in, or use a different SSO ID."

    users[sso_key] = {
        "name": name.strip(), "email": email.strip(), "password_hash": _hash_password(password),
        "role": ROLE_USER, "apps": [], "active": True, "status": STATUS_PENDING,
        "created_on": now_ist().isoformat(),
    }
    save_users(users)
    return True, "Your account request has been submitted and sent to the Admin for approval."


def change_password(sso_id: str, current_password: str, new_password: str):
    if not new_password or len(new_password) < 15:
        return False, "New password must be at least 15 characters."
    sso_key = sso_id.strip().lower()
    users = load_users()
    user = users.get(sso_key)
    if user is None:
        return False, "User not found."
    if user.get("password_hash") != _hash_password(current_password):
        return False, "Current password is incorrect."
    user["password_hash"] = _hash_password(new_password)
    save_users(users)
    return True, "Password updated successfully."


def add_or_update_user(sso_id: str, name: str, email: str, role: str, apps: list,
                        password: str = None, active: bool = True, status: str = None):
    sso_key = sso_id.strip().lower()
    users = load_users()
    existing = users.get(sso_key, {})
    password_hash = existing.get("password_hash")
    if password:
        password_hash = _hash_password(password)
    if not password_hash:
        password_hash = _hash_password("Welcome@123")
    users[sso_key] = {
        "name": name.strip(), "email": email.strip(), "password_hash": password_hash,
        "role": role, "apps": [a for a in apps if a in ALL_APPS], "active": active,
        "status": status if status else existing.get("status", STATUS_APPROVED),
        "created_on": existing.get("created_on", now_ist().isoformat()),
    }
    save_users(users)
    return True


def approve_user(sso_id: str, role: str, apps: list):
    sso_key = sso_id.strip().lower()
    users = load_users()
    if sso_key not in users:
        return False
    users[sso_key]["role"] = role
    users[sso_key]["apps"] = [a for a in apps if a in ALL_APPS]
    users[sso_key]["status"] = STATUS_APPROVED
    users[sso_key]["active"] = True
    save_users(users)
    return True


def reject_user(sso_id: str):
    sso_key = sso_id.strip().lower()
    users = load_users()
    if sso_key not in users:
        return False
    users[sso_key]["status"] = STATUS_REJECTED
    users[sso_key]["apps"] = []
    save_users(users)
    return True


def delete_user(sso_id: str):
    sso_key = sso_id.strip().lower()
    users = load_users()
    if sso_key in users:
        del users[sso_key]
        save_users(users)
        return True
    return False


def set_user_active(sso_id: str, active: bool):
    sso_key = sso_id.strip().lower()
    users = load_users()
    if sso_key in users:
        users[sso_key]["active"] = active
        save_users(users)
        return True
    return False


# ============================================================================
# SECTION 2 — THEME
# ============================================================================
PRIMARY_PURPLE, PRIMARY_PURPLE_DARK, PRIMARY_PURPLE_DARKER = "#5B2A86", "#3B1D57", "#2A1440"
PRIMARY_PURPLE_LIGHT, PRIMARY_PURPLE_SOFT, ACCENT_LILAC = "#F3EAFB", "#EFE3FA", "#8E5FB5"
OFF_WHITE, TEXT_DARK, TEXT_MUTED, BORDER_SOFT = "#FAFAFC", "#1F1B24", "#6B6575", "#E4D9F0"


def inject_global_css():
    st.markdown(f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
        * {{ box-sizing: border-box; }}
        html, body, [class*="css"] {{ font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important; }}
        .stApp {{ background: linear-gradient(180deg, {OFF_WHITE} 0%, #FFF 260px); }}
        header[data-testid="stHeader"] {{ background-color: {PRIMARY_PURPLE}; box-shadow: 0 2px 8px rgba(91,42,134,0.25); }}
        .block-container {{ padding-top: 2.6rem !important; padding-bottom: 3rem !important; max-width: 1180px; }}
        section[data-testid="stSidebar"] {{ background: linear-gradient(180deg, {PRIMARY_PURPLE_DARKER}, {PRIMARY_PURPLE_DARK}); }}
        section[data-testid="stSidebar"] > div:first-child {{ padding-top: 1.4rem; }}
        section[data-testid="stSidebar"] * {{ color: #F3EEFA !important; }}
        section[data-testid="stSidebar"] hr {{ border-top: 1px solid rgba(255,255,255,0.14) !important; margin: 1rem 0 !important; }}
        .sidebar-user-card {{ background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.12); border-radius: 14px; padding: 18px 16px 16px 16px; margin-bottom: 6px; text-align: center; display: flex; flex-direction: column; align-items: center; }}
        .sidebar-avatar {{ width: 54px; height: 54px; border-radius: 50%; background: linear-gradient(135deg, {ACCENT_LILAC}, {PRIMARY_PURPLE}); display: flex; align-items: center; justify-content: center; font-weight: 800; font-size: 20px; color: #FFF; margin-bottom: 10px; border: 2px solid rgba(255,255,255,0.35); flex-shrink: 0; }}
        .sidebar-name {{ font-weight: 700 !important; font-size: 15px !important; color: #FFF !important; margin-bottom: 3px; }}
        .sidebar-sso {{ font-size: 11.5px !important; color: #D8C7EC !important; font-family: 'Courier New', monospace; margin-bottom: 8px; }}
        .sidebar-role-pill {{ display: inline-block; background: #FFF !important; color: {PRIMARY_PURPLE_DARK} !important; border-radius: 999px; padding: 4px 14px; font-size: 11.5px !important; font-weight: 700 !important; }}
        section[data-testid="stSidebar"] .sidebar-role-pill {{ color: {PRIMARY_PURPLE_DARK} !important; background: #FFFFFF !important; }}
        section[data-testid="stSidebar"] .stRadio label {{ color: #F3EEFA !important; font-size: 14px !important; }}
        section[data-testid="stSidebar"] div[data-testid="stExpander"] {{ background: rgba(255,255,255,0.06) !important; border: 1px solid rgba(255,255,255,0.16) !important; }}
        section[data-testid="stSidebar"] div[data-testid="stExpander"] summary {{ background: transparent !important; color: #FFF !important; font-weight: 600 !important; }}
        section[data-testid="stSidebar"] div[data-testid="stExpander"] * {{ color: #F3EEFA !important; }}
        section[data-testid="stSidebar"] .stTextInput input {{ background: rgba(255,255,255,0.9) !important; color: #1F1B24 !important; }}
        h1, h2, h3 {{ color: {PRIMARY_PURPLE_DARK} !important; font-weight: 700 !important; letter-spacing: -0.2px; }}
        h3 {{ font-size: 1.22rem !important; margin-top: 0.2rem !important; }}
        .stCaption, .stMarkdown p, label, .stTextInput label, .stTextArea label, .stSelectbox label, .stMultiSelect label, .stRadio label, .stCheckbox label {{ color: {TEXT_DARK} !important; }}
        .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] > div {{ background: #FFF !important; color: {TEXT_DARK} !important; border: 1.5px solid {BORDER_SOFT} !important; border-radius: 10px !important; }}
        .stTextInput input:focus, .stTextArea textarea:focus {{ border: 1.5px solid {PRIMARY_PURPLE} !important; box-shadow: 0 0 0 3px rgba(91,42,134,0.12) !important; }}
        .stTextInput input::placeholder, .stTextArea textarea::placeholder {{ color: #9B93A8 !important; opacity: 1 !important; }}
        .stButton > button, div[data-testid="stFormSubmitButton"] button {{ background: linear-gradient(135deg, {PRIMARY_PURPLE}, {PRIMARY_PURPLE_DARK}) !important; color: #FFF !important; border: none !important; border-radius: 10px !important; font-weight: 600 !important; padding: 0.55rem 1.2rem !important; width: 100%; box-shadow: 0 2px 6px rgba(91,42,134,0.25); transition: all 0.15s ease-in-out; min-height: 42px; }}
        .stButton > button:hover, div[data-testid="stFormSubmitButton"] button:hover {{ box-shadow: 0 4px 12px rgba(91,42,134,0.35); transform: translateY(-1px); }}
        .stDownloadButton > button {{ background: #FFF !important; color: {PRIMARY_PURPLE_DARK} !important; border: 1.5px solid {PRIMARY_PURPLE} !important; border-radius: 10px !important; font-weight: 600 !important; }}
        div[data-testid="column"] .stButton > button {{ width: 100%; }}
        div[data-baseweb="tab-list"] {{ gap: 6px; border-bottom: 1px solid {BORDER_SOFT}; }}
        button[data-baseweb="tab"] {{ font-weight: 600 !important; color: {TEXT_MUTED} !important; font-size: 14.5px !important; padding: 10px 6px !important; }}
        button[data-baseweb="tab"][aria-selected="true"] {{ color: {PRIMARY_PURPLE_DARK} !important; }}
        div[data-baseweb="tab-highlight"] {{ background-color: {PRIMARY_PURPLE} !important; height: 3px !important; }}
        div[data-testid="stAlert"] {{ border: none !important; border-left: 4px solid {PRIMARY_PURPLE} !important; border-radius: 10px !important; background: {PRIMARY_PURPLE_SOFT} !important; color: {TEXT_DARK} !important; }}
        div[data-testid="stExpander"] {{ border: 1px solid {BORDER_SOFT} !important; border-radius: 12px !important; background: #FFF !important; box-shadow: 0 1px 4px rgba(91,42,134,0.05); overflow: hidden; }}
        div[data-testid="stExpander"] summary {{ background: {PRIMARY_PURPLE_SOFT} !important; color: {PRIMARY_PURPLE_DARK} !important; font-weight: 600 !important; }}
        div[data-testid="stExpander"] div[data-testid="stExpanderDetails"], div[data-testid="stExpander"] div[data-testid="stExpanderDetails"] p, div[data-testid="stExpander"] div[data-testid="stExpanderDetails"] li {{ color: {TEXT_DARK} !important; }}
        div[data-testid="stVerticalBlockBorderWrapper"] {{ background: #FFF !important; border: 1px solid {BORDER_SOFT} !important; border-radius: 18px !important; box-shadow: 0 8px 28px rgba(91,42,134,0.10) !important; padding: 0.3rem 0.3rem !important; }}
        div[data-testid="stDataFrame"] {{ border: 1px solid {BORDER_SOFT} !important; border-radius: 12px !important; overflow: hidden; }}
        div[data-testid="stMetric"] {{ background: #FFF; border: 1px solid {BORDER_SOFT}; border-radius: 12px; padding: 14px 16px 10px 16px; box-shadow: 0 1px 4px rgba(91,42,134,0.05); text-align: center; }}
        div[data-testid="stMetricLabel"] {{ color: {TEXT_MUTED} !important; font-weight: 600 !important; font-size: 12px !important; text-transform: uppercase; justify-content: center !important; }}
        div[data-testid="stMetricValue"] {{ color: {PRIMARY_PURPLE_DARK} !important; font-weight: 800 !important; justify-content: center !important; }}
        hr {{ border: none !important; border-top: 1px solid {BORDER_SOFT} !important; margin: 1.3rem 0 !important; }}
        .portal-header-banner {{ background: linear-gradient(120deg, {PRIMARY_PURPLE}, {PRIMARY_PURPLE_DARK}); color: #FFF; border-radius: 16px; padding: 24px 28px; margin-bottom: 22px; display: flex; align-items: center; gap: 16px; box-shadow: 0 6px 20px rgba(91,42,134,0.28); }}
        .portal-header-icon {{ width: 50px; height: 50px; border-radius: 14px; background: rgba(255,255,255,0.16); border: 1px solid rgba(255,255,255,0.35); display: flex; align-items: center; justify-content: center; font-weight: 800; font-size: 19px; color: #FFF; flex-shrink: 0; }}
        /* Scoped fix: only columns that directly wrap a dashboard .app-tile card
           get stretched to equal height -- this no longer affects file-uploader
           columns, the tool top-bar columns, or any other column layout in the
           app (which was the root cause of the misalignment seen previously). */
        div[data-testid="column"]:has(.app-tile) {{ display: flex; }}
        div[data-testid="column"]:has(.app-tile) > div {{ width: 100%; }}
        .app-tile {{ background: #FFF; border: 1px solid {BORDER_SOFT}; border-radius: 16px; padding: 24px 20px; text-align: center; box-shadow: 0 3px 12px rgba(91,42,134,0.07); transition: all 0.18s ease-in-out; display: flex; flex-direction: column; align-items: center; height: 100%; min-height: 210px; }}

        /* Normalize widget label height so columns with a longer (2-line) label
           next to columns with a short (1-line) label still start their input/
           button at the exact same vertical position -- fixes misaligned
           multi-column file-uploader rows (e.g. in the Saviynt tool) without
           touching that tool's own code. */
        div[data-testid="stWidgetLabel"] {{ min-height: 44px; display: flex; align-items: flex-end; }}
        div[data-testid="stFileUploader"] section {{ margin-top: 0 !important; }}
        .app-tile:hover {{ box-shadow: 0 10px 26px rgba(91,42,134,0.18); transform: translateY(-2px); border-color: {PRIMARY_PURPLE}; }}
        .app-tile-icon-badge {{ width: 54px; height: 54px; border-radius: 14px; background: linear-gradient(135deg, {PRIMARY_PURPLE_SOFT}, #FFF); border: 1px solid {BORDER_SOFT}; display: flex; align-items: center; justify-content: center; font-size: 25px; margin-bottom: 12px; flex-shrink: 0; }}
        .app-tile-title {{ font-weight: 700; font-size: 16px; color: {TEXT_DARK}; margin-bottom: 8px; }}
        .app-tile-desc {{ font-size: 13px; color: {TEXT_MUTED}; line-height: 1.5; flex-grow: 1; }}
        .role-badge {{ display: inline-block; background: {PRIMARY_PURPLE_LIGHT} !important; color: {PRIMARY_PURPLE_DARK} !important; border: 1px solid {PRIMARY_PURPLE}; border-radius: 999px; padding: 3px 14px; font-size: 12px; font-weight: 700; }}
        .breadcrumb-text {{ font-size: 14px; color: {TEXT_MUTED}; padding-top: 10px; }}
        .breadcrumb-text .current {{ color: {PRIMARY_PURPLE_DARK}; font-weight: 700; }}
        .access-denied-card {{ background: #FFF5F5; border-left: 5px solid #D64545; border-radius: 14px; padding: 22px 26px; margin: 10px 0 20px 0; }}
        .pending-card {{ background: {PRIMARY_PURPLE_SOFT}; border-left: 5px solid {PRIMARY_PURPLE}; border-radius: 14px; padding: 22px 26px; margin: 10px 0 20px 0; }}
        .request-card {{ border: 1px solid {BORDER_SOFT}; border-radius: 14px; padding: 16px 18px; margin-bottom: 14px; background: #FFFFFF; }}
        .pending-badge {{ display: inline-block; background: #D64545; color: #FFF; border-radius: 999px; font-size: 10.5px; font-weight: 700; padding: 1px 7px; margin-left: 6px; vertical-align: middle; }}
        </style>
        """, unsafe_allow_html=True)


def render_header_banner(title: str, subtitle: str = ""):
    st.markdown(f"""
        <div class="portal-header-banner">
            <div class="portal-header-icon">GE</div>
            <div>
                <div style="color:#E4D2F5;font-size:12px;letter-spacing:1.1px;text-transform:uppercase;font-weight:600;">{subtitle}</div>
                <div style="color:#FFF;font-size:24px;font-weight:800;line-height:1.3;">{title}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ============================================================================
# SECTION 3 — LOGIN PAGE (Sign In + Create Account)
# ============================================================================

def _render_sign_in_tab():
    with st.form("login_form", clear_on_submit=False):
        sso_id = st.text_input("SSO ID", key="login_sso_input")
        password = st.text_input("Password", type="password", placeholder="Enter your password", key="login_pwd_input")
        submitted = st.form_submit_button("Sign In", use_container_width=True)

    if submitted:
        ok, result = authenticate(sso_id, password)
        if ok:
            st.session_state.authenticated = True
            st.session_state.current_user = result
            st.session_state.plain_password = password
            st.session_state.login_error = ""
            st.rerun()
        else:
            st.session_state.login_error = result

    if st.session_state.get("login_error"):
        st.error(st.session_state.login_error)
        st.caption("Please re-enter your correct SSO ID and Password and try again.")


def _render_create_account_tab():
    st.markdown(f"<p style='color:{TEXT_MUTED};font-size:13px;margin-top:-6px;margin-bottom:14px;'>Create your SSO account below. Your request will be sent to the Admin for approval before you can access any application.</p>", unsafe_allow_html=True)
    with st.form("register_form", clear_on_submit=False):
        reg_sso_id = st.text_input("Choose an SSO ID *", key="reg_sso_input")
        reg_name = st.text_input("Full Name *", key="reg_name_input")
        reg_email = st.text_input("Email *", key="reg_email_input")
        reg_password = st.text_input("Password *", type="password", placeholder="At least 15 characters", key="reg_pwd_input")
        reg_password_confirm = st.text_input("Confirm Password *", type="password", key="reg_pwd_confirm_input")
        reg_submitted = st.form_submit_button("Create Account", use_container_width=True)

    if reg_submitted:
        if reg_password != reg_password_confirm:
            st.session_state.register_error = "Passwords do not match. Please re-enter."
            st.session_state.register_success = ""
        else:
            ok, msg = register_user(reg_sso_id, reg_name, reg_email, reg_password)
            if ok:
                st.session_state.register_success = msg
                st.session_state.register_error = ""
            else:
                st.session_state.register_error = msg
                st.session_state.register_success = ""

    if st.session_state.get("register_error"):
        st.error(st.session_state.register_error)
    if st.session_state.get("register_success"):
        st.success(f"✅ {st.session_state.register_success}")
        st.info("You may sign in now to check your approval status at any time from the **Sign In** tab.")


def render_login_page():
    if "register_error" not in st.session_state:
        st.session_state.register_error = ""
    if "register_success" not in st.session_state:
        st.session_state.register_success = ""

    st.markdown("<div style='height:2rem;'></div>", unsafe_allow_html=True)
    col_l, col_mid, col_r = st.columns([1, 1.3, 1])
    with col_mid:
        st.markdown(f"""
            <div style="text-align:center;margin-bottom:20px;">
                <div style="width:62px;height:62px;border-radius:16px;margin:0 auto 12px auto;background:linear-gradient(135deg,{PRIMARY_PURPLE},{PRIMARY_PURPLE_DARK});display:flex;align-items:center;justify-content:center;font-weight:800;font-size:23px;color:#FFF;box-shadow:0 8px 20px rgba(91,42,134,0.30);">GE</div>
                <div style="font-size:21px;font-weight:800;color:{PRIMARY_PURPLE_DARK};">Tech Assisted Audit Tools</div>
                <div style="font-size:12.5px;color:{TEXT_MUTED};text-transform:uppercase;letter-spacing:0.4px;margin-top:2px;">GEHC IT SOX Team</div>
            </div>
            """, unsafe_allow_html=True)

        with st.container(border=True):
            tab_signin, tab_register = st.tabs(["🔐 Sign In", "📝 Create Account"])
            with tab_signin:
                st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
                _render_sign_in_tab()
            with tab_register:
                st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
                _render_create_account_tab()

        with st.expander("How to Access and Use the Tool"):
            st.markdown(
                "- If you already have an account, simply log in to start using the platform.\n"
                "- If you do not have an account, create one and submit your registration request. An administrator will review and approve your request. Once approved, you can log in and access the platform.\n"
                "- After logging in, select the tool you want to use and get started."
            )


# ============================================================================
# SECTION 4 — ADMIN PORTAL (Admin role ONLY)
# ============================================================================

def _users_to_dataframe(users: dict) -> pd.DataFrame:
    rows = [{"SSO ID": sid, "Name": u.get("name",""), "Email": u.get("email",""),
             "Role": u.get("role",""), "Apps": ", ".join(u.get("apps",[])),
             "Status": u.get("status", STATUS_APPROVED).capitalize(),
             "Active": u.get("active", True)}
            for sid, u in users.items()]
    cols = ["SSO ID","Name","Email","Role","Apps","Status","Active"]
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)


def _render_pending_requests_section(users: dict):
    pending = {sid: u for sid, u in users.items() if u.get("status") == STATUS_PENDING}
    badge = f"<span class='pending-badge'>{len(pending)} NEW</span>" if pending else ""
    st.markdown(f"<h3>🔔 Pending Approval Requests {badge}</h3>", unsafe_allow_html=True)

    if not pending:
        st.info("No pending account requests right now.")
        return

    for sid, u in pending.items():
        with st.container(border=True):
            col_info, col_meta = st.columns([3, 2])
            with col_info:
                st.markdown(f"**{u.get('name','')}**  \n`{sid}` &nbsp;·&nbsp; {u.get('email','')}")
            with col_meta:
                st.caption(f"Requested on {str(u.get('created_on',''))[:10]}")

            with st.form(f"approve_form_{sid}"):
                fc1, fc2 = st.columns(2)
                with fc1:
                    assign_role = st.selectbox("Assign Role", options=ALL_ROLES, index=ALL_ROLES.index(ROLE_USER), key=f"role_select_{sid}")
                with fc2:
                    assign_apps = st.multiselect("Assign Application Access", options=ALL_APPS, key=f"apps_select_{sid}")
                bc1, bc2 = st.columns(2)
                with bc1:
                    approve_clicked = st.form_submit_button("✅ Approve", use_container_width=True)
                with bc2:
                    reject_clicked = st.form_submit_button("🚫 Reject", use_container_width=True)

            if approve_clicked:
                approve_user(sid, assign_role, assign_apps)
                st.success(f"✅ `{sid}` approved with role '{assign_role}' and access to: {', '.join(assign_apps) or '(no apps selected)'}.")
                st.rerun()
            if reject_clicked:
                reject_user(sid)
                st.warning(f"🚫 `{sid}` has been rejected.")
                st.rerun()

    st.markdown("---")


def render_admin_portal():
    current_user = st.session_state.current_user
    current_role = current_user.get("role")

    if current_role != ROLE_ADMIN:
        render_header_banner("Access Restricted", "Admin Portal")
        st.markdown(f"""
            <div class="access-denied-card">
                <div style="font-size:17px;font-weight:700;color:#B33A3A;">🚫 You don't have permission to view this page</div>
                <div style="color:{TEXT_MUTED};font-size:14px;">The Admin Portal is restricted to <b>Administrator</b> accounts only. Your current role is <b>{current_role}</b>.</div>
            </div>
            """, unsafe_allow_html=True)
        st.session_state.portal_view = "dashboard"
        if st.button("⬅ Return to My Dashboard"):
            st.rerun()
        return

    render_header_banner("Admin Portal", "User & Access Management")
    st.markdown(f"Signed in as **{current_user.get('name')}** (<span class='role-badge'>{current_role}</span>)", unsafe_allow_html=True)
    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    users = load_users()
    _render_pending_requests_section(users)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Users", len(users))
    m2.metric("Admins", sum(1 for u in users.values() if u.get("role") == ROLE_ADMIN))
    m3.metric("SOX Team", sum(1 for u in users.values() if u.get("role") == ROLE_SOX))
    m4.metric("Rejected User", sum(1 for u in users.values() if u.get("role") == ROLE_USER))
    m5.metric("Active Accounts", sum(1 for u in users.values() if u.get("active", True)))

    st.markdown("---")
    st.subheader("👥 User Directory")
    st.dataframe(_users_to_dataframe(users), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("➕ Add / Update User or Admin")
    _render_admin_add_user_form(users)

    st.markdown("---")
    st.subheader("⚙️ Manage Existing Users")
    _render_manage_existing_users(users, current_role)


def _render_admin_add_user_form(users: dict):
    with st.form("admin_add_user_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            sso_id = st.text_input("SSO ID *", placeholder="e.g. jsmith45")
            name = st.text_input("Full Name *", placeholder="e.g. John Smith")
        with col2:
            email = st.text_input("Email *", placeholder="e.g. jsmith45@gehealthcare.com")
            role = st.selectbox("Role *", options=ALL_ROLES)
        apps = st.multiselect("Application Access (multiple allowed)", options=ALL_APPS)
        password = st.text_input("Temporary Password (leave blank for default)", type="password")
        active = st.checkbox("Active", value=True)
        submitted = st.form_submit_button("Save User", use_container_width=True)
    if submitted:
        if not sso_id or not name or not email:
            st.warning("Please fill in SSO ID, Name, and Email.")
        else:
            add_or_update_user(sso_id=sso_id, name=name, email=email, role=role, apps=apps, password=password if password else None, active=active, status=STATUS_APPROVED)
            st.success(f"User `{sso_id.strip().lower()}` saved successfully with role '{role}'.")
            st.rerun()


def _render_manage_existing_users(users: dict, current_role: str):
    if not users:
        st.info("No users in the directory yet.")
        return
    selected_sso = st.selectbox("Select a user to manage", options=list(users.keys()), key="manage_user_select")
    user = users[selected_sso]

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"**Name:** {user.get('name')}")
        st.markdown(f"**Email:** {user.get('email')}")
    with col2:
        st.markdown(f"**Role:** {user.get('role')}")
        st.markdown(f"**Apps:** {', '.join(user.get('apps', [])) or '(none)'}")
    with col3:
        status_val = user.get("status", STATUS_APPROVED)
        status_icon = {"approved": "🟢", "pending": "🟡", "rejected": "🔴"}.get(status_val, "⚪")
        st.markdown(f"**Status:** {status_icon} {status_val.capitalize()}")

    with st.expander("✏️ Edit this user", expanded=False):
        with st.form(f"edit_user_form_{selected_sso}"):
            new_name = st.text_input("Full Name", value=user.get("name", ""))
            new_email = st.text_input("Email", value=user.get("email", ""))
            new_role = st.selectbox("Role", options=ALL_ROLES, index=ALL_ROLES.index(user.get("role")) if user.get("role") in ALL_ROLES else 0)
            new_apps = st.multiselect("Application Access", options=ALL_APPS, default=user.get("apps", []))
            new_password = st.text_input("Reset Password (leave blank to keep current)", type="password")
            save_edit = st.form_submit_button("Save Changes", use_container_width=True)
        if save_edit:
            add_or_update_user(sso_id=selected_sso, name=new_name, email=new_email, role=new_role, apps=new_apps, password=new_password if new_password else None, active=user.get("active", True), status=user.get("status", STATUS_APPROVED))
            st.success(f"User `{selected_sso}` updated successfully.")
            st.rerun()

    colA, colB = st.columns(2)
    with colA:
        if user.get("active", True):
            if st.button("🔒 Deactivate User", key=f"deact_{selected_sso}"):
                set_user_active(selected_sso, False)
                st.success(f"User `{selected_sso}` deactivated.")
                st.rerun()
        else:
            if st.button("🔓 Reactivate User", key=f"react_{selected_sso}"):
                set_user_active(selected_sso, True)
                st.success(f"User `{selected_sso}` reactivated.")
                st.rerun()
    with colB:
        if st.button("🗑️ Delete User", key=f"del_{selected_sso}"):
            if selected_sso == st.session_state.current_user.get("sso_id"):
                st.error("You cannot delete your own currently logged-in account.")
            else:
                delete_user(selected_sso)
                st.success(f"User `{selected_sso}` deleted.")
                st.rerun()


# ============================================================================
# SECTION 5 — APPLICATION DASHBOARD
# ============================================================================
APP_ICONS = {APP_SAVIYNT: "🛡️", APP_CM_AUTOMATION: "⚙️"}
APP_DESCRIPTIONS = {
    APP_SAVIYNT: "App provisioning testing: Validate access requests for Saviynt based approvals, roles provisioned and export the results in an excel report.",
    APP_CM_AUTOMATION: "Change Management SOX testing: Parse change tickets (PDF), perform IT SOD check, validate CAB approvals and export the results in an excel report. ",
}


def render_pending_screen():
    user = st.session_state.current_user
    render_header_banner("Approval Pending", f"Welcome, {user.get('name')}")
    st.markdown(f"""
        <div class="pending-card">
            <div style="font-size:17px;font-weight:700;color:{PRIMARY_PURPLE_DARK};margin-bottom:6px;">⏳ Your account request has been sent to the Admin for approval</div>
            <div style="color:{TEXT_MUTED};font-size:14px;">An Administrator needs to review your sign-up, assign you a role, and grant application access before you can use any tool. You'll get full access automatically once approved -- no need to sign up again.</div>
        </div>
        """, unsafe_allow_html=True)
    if st.button("🔄 Check Approval Status"):
        st.rerun()


def _render_tool_top_bar(active_tool: str):
    st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
    col_back, col_crumb = st.columns([1.1, 4], gap="medium")
    with col_back:
        if st.button("⬅ Back to Dashboard", key="back_to_dashboard_top", use_container_width=True):
            st.session_state.active_tool = None
            st.rerun()
    with col_crumb:
        st.markdown(
            f"""
            <div style="display:flex;align-items:center;height:42px;">
                <span class="breadcrumb-text" style="padding-top:0;">
                    Application Dashboard &nbsp;/&nbsp; <span class="current">{active_tool}</span>
                </span>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.markdown("---")


def _render_tool_page(active_tool: str):
    _render_tool_top_bar(active_tool)
    render_header_banner(active_tool, "Application")
    st.markdown("---")
    if active_tool == APP_SAVIYNT:
        render_saviynt_tool()
    elif active_tool == APP_CM_AUTOMATION:
        render_cm_automation_tool()


def render_dashboard():
    user = st.session_state.current_user
    user_apps = [a for a in user.get("apps", []) if a in ALL_APPS]
    active_tool = st.session_state.get("active_tool")

    if active_tool and active_tool in user_apps:
        _render_tool_page(active_tool)
        return

    render_header_banner("Application Dashboard", f"Welcome, {user.get('name')}")
    st.markdown(f"Role: <span class='role-badge'>{user.get('role')}</span>", unsafe_allow_html=True)
    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

    if not user_apps:
        st.warning("You currently have no application access assigned. Please contact your Admin or the GEHC IT SOX Team to request access.")
        return

    st.subheader("Your Applications")
    cols = st.columns(2) if len(user_apps) <= 2 else st.columns(3)
    for i, app_name in enumerate(user_apps):
        col = cols[i % len(cols)]
        with col:
            st.markdown(f"""
                <div class="app-tile">
                    <div class="app-tile-icon-badge">{APP_ICONS.get(app_name, "📦")}</div>
                    <div class="app-tile-title">{app_name}</div>
                    <div class="app-tile-desc">{APP_DESCRIPTIONS.get(app_name, "")}</div>
                </div>
                """, unsafe_allow_html=True)
            st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
            if st.button(f"Open {app_name}", key=f"open_{app_name}", use_container_width=True):
                st.session_state.active_tool = app_name
                st.rerun()


# ============================================================================
# SECTION 6: SAVIYNT TOOL (SOX Access Comparison Tool)
# ============================================================================


def render_saviynt_tool():
    st.markdown("### 🛡️ SOX Access Comparison Tool (Saviynt)")
    st.caption("Upload the Analytics Summary and Dump files, run checks, and see per‑check summaries.")


    # =========================================
    #  Helper functions
    # =========================================
    def make_unique(cols):
        """Make column names unique (for Streamlit/pyarrow display)."""
        seen = {}
        new_cols = []
        for c in cols:
            if c not in seen:
                seen[c] = 0
                new_cols.append(c)
            else:
                seen[c] += 1
                new_cols.append(f"{c}_{seen[c]}")
        return new_cols


    def extract_id_from_username(username: str) -> str:
        """
        Extract ID from a string like 'Name (12345)' => '12345'.
        Returns upper-stripped string or '' if not found.
        """
        if pd.isna(username):
            return ""
        text = str(username)
        match = re.search(r"\(([^)]+)\)", text)
        if match:
            return match.group(1).strip().upper()
        return ""


    def parse_multi_values(value, treat_as_name_id=False):
        """
        Split a cell into a set of normalized approver/owner values.
        Handles comma-separated values like 'a, b, c'
        or 'Name (ID), Name2 (ID2)' — even if names themselves contain commas
        like 'Smith, John (12345), Doe, Jane (67890)'.

        If treat_as_name_id=True:
            - Splits on commas ONLY when they are NOT inside parentheses
              (so 'Smith, John (12345)' stays as one chunk before ID extraction).
            - Extracts the ID from inside brackets for each chunk.

        If treat_as_name_id=False:
            - Splits on every comma (simple case, plain IDs/names).
        """
        if pd.isna(value):
            return set()
        text = str(value)

        if treat_as_name_id:
            # Split on comma only if NOT followed eventually by ')' before another '('
            parts = re.split(r',\s*(?![^()]*\))', text)
        else:
            parts = text.split(",")

        parts = [p.strip() for p in parts if p.strip()]

        result = set()
        for p in parts:
            if treat_as_name_id:
                id_ = extract_id_from_username(p)
                result.add(id_ if id_ else p.strip().upper())
            else:
                result.add(p.strip().upper())
        return result


    def build_role_approver_dict(df, role_col, approver_cols, treat_as_name_id=False):
        """
        Build {role: set(approvers)} from a dataframe (used as a reference/lookup table).
        approver_cols: list of column names (values may be comma-separated).
        Duplicate roles across rows are merged (union of approvers).
        """
        result = {}
        for _, row in df.iterrows():
            role_val = row[role_col]
            if pd.isna(role_val) or str(role_val).strip() == "":
                continue
            role = str(role_val).strip().upper()

            approvers = set()
            for col in approver_cols:
                approvers |= parse_multi_values(row[col], treat_as_name_id)

            if role in result:
                result[role] |= approvers
            else:
                result[role] = approvers
        return result


    def format_approval_message(role_display, common_approvers):
        """
        Build the human-readable message:
        - 0 matches  -> 'has no matching approvers'
        - 1 match    -> 'is approved by X'
        - 2 matches  -> 'is approved by both X and Y'
        - 3+ matches -> 'is approved by X, Y, and Z'
        """
        approvers = sorted(common_approvers)
        if len(approvers) == 0:
            return f"{role_display} has no matching approvers"
        elif len(approvers) == 1:
            return f"{role_display} is approved by {approvers[0]}"
        elif len(approvers) == 2:
            return f"{role_display} is approved by both {approvers[0]} and {approvers[1]}"
        else:
            return f"{role_display} is approved by {', '.join(approvers[:-1])}, and {approvers[-1]}"


    def style_dataframe_safe(df, status_col=None, valid_cols=None,
                              status_style_fn=None, valid_style_fn=None):
        """
        Apply cell-level styling to a dataframe using pandas Styler,
        compatible with both old pandas (.applymap) and new pandas (.map).
        Requires df to have unique column names.
        """
        if df.columns.duplicated().any():
            raise ValueError("Cannot style a dataframe with duplicate column names.")

        styled = df.style

        use_map = hasattr(styled, "map")

        if status_col and status_col in df.columns and status_style_fn:
            if use_map:
                styled = styled.map(status_style_fn, subset=[status_col])
            else:
                styled = styled.applymap(status_style_fn, subset=[status_col])

        if valid_cols and valid_style_fn:
            valid_cols_present = [c for c in valid_cols if c in df.columns]
            if valid_cols_present:
                if use_map:
                    styled = styled.map(valid_style_fn, subset=valid_cols_present)
                else:
                    styled = styled.applymap(valid_style_fn, subset=valid_cols_present)

        return styled


    # =========================================
    #  1. File upload (all files used across checks)
    # =========================================
    st.header("Upload Files")

    col_up1, col_up2, col_up3 = st.columns(3)

    with col_up1:
        analytics_file = st.file_uploader(
            "Upload Saviynt Summary file (e.g. Analytics_Summary_*.xlsx)",
            type=["xlsx"],
            key="analytics"
        )

    with col_up2:
        sbl_file = st.file_uploader(
            "Upload Dump Automated file (e.g. sbl_am_automated.xlsx)",
            type=["xlsx"],
            key="sbl"
        )

    with col_up3:
        entitlement_file = st.file_uploader(
            "Upload Entitlement Owner List (optional, for Check 4)",
            type=["xlsx", "csv"],
            key="entitlement"
        )

    df_analytics, df_sbl, df_entitlement = None, None, None

    if analytics_file is not None:
        df_analytics = pd.read_excel(analytics_file)

    if sbl_file is not None:
        df_sbl = pd.read_excel(sbl_file)

    if entitlement_file is not None:
        if entitlement_file.name.endswith(".csv"):
            df_entitlement = pd.read_csv(entitlement_file)
        else:
            df_entitlement = pd.read_excel(entitlement_file)

    # Extra safety: stop execution until both mandatory files are uploaded
    if df_analytics is None or df_sbl is None:
        st.info("Please upload both files (Saviynt Summary and Dump Automated) to continue.")
        st.stop()

    # Row counts summary (now guaranteed not None)
    analytics_rows = len(df_analytics)
    sbl_rows = len(df_sbl)
    entitlement_msg = (
        f" Entitlement Owner file has {len(df_entitlement)} rows."
        if df_entitlement is not None
        else " Entitlement Owner file not uploaded (Check 4 will be skipped)."
    )
    st.success(
        f"Files uploaded successfully. "
        f"Saviynt file has {analytics_rows} rows. "
        f"Dump Automated file has {sbl_rows} rows."
        f"{entitlement_msg}"
    )

    all_results = []  # collect partial results from all checks

    # =========================================
    #  2. Check 1 – Date comparisons (Mode A & Mode B)
    # =========================================
    st.header("Check 1 – Access is provisioned post approvals only")

    check1_mode = st.radio(
        "Choose Check 1 mode:",
        options=[
            "A: Approval vs Task Completion (Saviynt only)",
            "B: Application Created Date vs Saviynt Request Approved Date (Unique by SSO)"
        ],
        index=0,
        key="check1_mode"
    )

    # ----------------- MODE A: Existing logic -----------------
    if check1_mode.startswith("A"):
        with st.expander("Configure Check 1 (Mode A)", expanded=True):
            c1_1, c1_2, c1_3 = st.columns(3)

            with c1_1:
                approval_col = st.selectbox(
                    "Approval Date column (Analytics)",
                    options=df_analytics.columns,
                    key="approval_date_col"
                )

            with c1_2:
                completion_col = st.selectbox(
                    "Task Completion Date column (Analytics)",
                    options=df_analytics.columns,
                    key="completion_date_col"
                )

            with c1_3:
                key_col_1 = st.selectbox(
                    "Key/ID column (Analytics, for report)",
                    options=df_analytics.columns,
                    key="key_col_1"
                )

            run_check1 = st.checkbox("Run Check 1 (Mode A, ALL rows)", value=True, key="run_check1A")

        if run_check1:
            df_check1 = df_analytics.copy()

            df_check1["_approval_datetime"] = pd.to_datetime(df_check1[approval_col], errors="coerce")
            df_check1["_completion_datetime"] = pd.to_datetime(df_check1[completion_col], errors="coerce")

            def check_approval_before_completion(row):
                a = row["_approval_datetime"]
                c = row["_completion_datetime"]
                if pd.isna(a) or pd.isna(c):
                    return "Failed - Missing date"
                return "Passed" if a <= c else "Failed - Approval after completion"

            df_check1["Check1_Result"] = df_check1.apply(check_approval_before_completion, axis=1)
            df_check1["Check1_Reason"] = df_check1["Check1_Result"]

            all_results.append(df_check1[[key_col_1, "Check1_Result", "Check1_Reason"]])

            total_rows_c1 = len(df_check1)
            failed_rows_c1 = df_check1[df_check1["Check1_Result"].str.startswith("Failed")]
            passed_rows_c1 = total_rows_c1 - len(failed_rows_c1)

            st.subheader("Check 1 (Mode A) Summary")
            st.write(f"Total rows evaluated: {total_rows_c1}")
            st.write(f"Passed: {passed_rows_c1}")
            st.write(f"Failed: {len(failed_rows_c1)}")

            if len(failed_rows_c1) > 0:
                st.write("Failed rows for Check 1 (Mode A):")
                failed_view = failed_rows_c1[[key_col_1, approval_col, completion_col, "Check1_Result"]].copy()
                failed_view.columns = make_unique(failed_view.columns)
                st.dataframe(failed_view)
            else:
                st.success("There were no issues or failed cases for this check (Mode A).")

            st.write("Full Check 1 (Mode A) results (ALL rows):")
            preview_cols = [key_col_1, approval_col, completion_col, "Check1_Result"]
            preview = df_check1[preview_cols].copy()
            preview.columns = make_unique(preview.columns)
            st.dataframe(preview)

    # ----------------- MODE B: SBL Created Date vs Analytics Request Approved Date -----------------
    else:
        with st.expander("Configure Check 1 (Mode B)", expanded=True):
            c1b_1, c1b_2, c1b_3, c1b_4 = st.columns(4)

            with c1b_1:
                sbl_sso_col = st.selectbox(
                    "User SSO / ID column in Dump (Application file)",
                    options=df_sbl.columns,
                    key="sbl_sso_col_c1"
                )

            with c1b_2:
                sbl_created_date_col = st.selectbox(
                    "Created Date column in Dump (Application file)",
                    options=df_sbl.columns,
                    key="sbl_created_date_col_c1"
                )

            with c1b_3:
                analytics_sso_col = st.selectbox(
                    "User SSO / ID column in Saviynt file",
                    options=df_analytics.columns,
                    key="analytics_sso_col_c1"
                )

            with c1b_4:
                analytics_approved_date_col = st.selectbox(
                    "Request Approved Date column in Saviynt file",
                    options=df_analytics.columns,
                    key="analytics_approved_date_col_c1"
                )

            key_col_1b = st.selectbox(
                "Key/ID column (Saviynt, for report)",
                options=df_analytics.columns,
                key="key_col_1b"
            )

            username_is_pure_id_c1 = st.checkbox(
                "Saviynt SSO column already contains pure ID (no brackets)",
                value=True,
                key="username_is_pure_id_c1",
                help="Tick if the selected Saviynt user column is just the ID (e.g., 223055402), not 'Name (223055402)'."
            )

            run_check1B = st.checkbox(
                "Run Check 1 (Mode B, ALL rows)",
                value=True,
                key="run_check1B"
            )

        if run_check1B:
            # --- Normalize SSO/ID in SBL (Dump) ---
            sbl_tmp = df_sbl[[sbl_sso_col, sbl_created_date_col]].copy()

            # Ensure SSO column is a Series
            sso_series = sbl_tmp[sbl_sso_col]
            if isinstance(sso_series, pd.DataFrame):
                sso_series = sso_series.iloc[:, 0]

            sbl_tmp["_sso_norm"] = sso_series.astype(str).str.strip().str.upper()

            # Ensure created-date column is a Series
            created_series = sbl_tmp[sbl_created_date_col]
            if isinstance(created_series, pd.DataFrame):
                created_series = created_series.iloc[:, 0]

            # Convert created date to date (no time)
            sbl_tmp["_created_dt"] = pd.to_datetime(created_series, errors="coerce").dt.date

            # Map: SSO -> earliest created date (if multiple rows in SBL)
            sbl_created_map = (
                sbl_tmp.groupby("_sso_norm")["_created_dt"]
                .min()
                .to_dict()
            )

            # --- Normalize SSO/ID in Analytics ---
            df_check1B = df_analytics[[analytics_sso_col, analytics_approved_date_col, key_col_1b]].copy()

            analytics_user_series = df_check1B[analytics_sso_col]
            if isinstance(analytics_user_series, pd.DataFrame):
                analytics_user_series = analytics_user_series.iloc[:, 0]

            if username_is_pure_id_c1:
                df_check1B["_sso_norm"] = analytics_user_series.astype(str).str.strip().str.upper()
            else:
                df_check1B["_sso_norm"] = analytics_user_series.apply(extract_id_from_username)

            # Ensure approved-date column is a Series
            approved_series = df_check1B[analytics_approved_date_col]
            if isinstance(approved_series, pd.DataFrame):
                approved_series = approved_series.iloc[:, 0]

            # Convert approved date to date (no time)
            df_check1B["_approved_dt"] = pd.to_datetime(approved_series, errors="coerce").dt.date

            # --- Row-wise comparison: Approved Date vs Created Date ---
            def compare_approved_vs_created(row):
                sso = row["_sso_norm"]
                approved_dt = row["_approved_dt"]

                # If we can't extract any SSO at all from Analytics -> fail
                if not sso:
                    return "Failed - No SSO/ID extracted", None, ""

                created_dt = sbl_created_map.get(sso, None)

                # If SSO not found in SBL -> fail
                if created_dt is None:
                    return "Failed - SSO/ID not found in Application File", None, sso

                # If created date exists in mapping but is NaN -> fail
                if pd.isna(created_dt):
                    return "Failed - Created date missing in Application File", None, sso

                # From this point, we HAVE a created date in SBL
                if pd.isna(approved_dt):
                    return "Failed - Approved date missing in Saviynt File", created_dt, sso

                # Rule: Approved date must be <= Created date
                if approved_dt <= created_dt:
                    return "Passed", created_dt, sso
                else:
                    return "Failed - Approved date is after created date", created_dt, sso

            results_b = df_check1B.apply(
                lambda r: compare_approved_vs_created(r),
                axis=1,
                result_type="expand"
            )

            df_check1B["Check1_Result"] = results_b[0]
            df_check1B["Created Date"] = results_b[1]
            df_check1B["SSO ID"] = results_b[2]
            df_check1B["Approved Date"] = df_check1B["_approved_dt"]
            df_check1B["Check1_Reason"] = df_check1B["Check1_Result"]

            # --- Collect for final download / combined results ---
            all_results.append(
                df_check1B[
                    [
                        key_col_1b,
                        "SSO ID",
                        "Created Date",
                        "Approved Date",
                        "Check1_Result",
                        "Check1_Reason",
                    ]
                ]
            )

            total_rows_c1b = len(df_check1B)
            failed_rows_c1b = df_check1B[df_check1B["Check1_Result"].str.startswith("Failed")]
            passed_rows_c1b = total_rows_c1b - len(failed_rows_c1b)

            st.subheader("Check 1 (Mode B) Summary")
            st.write(f"Total rows evaluated (Analytics): {total_rows_c1b}")
            st.write(f"Passed: {passed_rows_c1b}")
            st.write(f"Failed: {len(failed_rows_c1b)}")

            if len(failed_rows_c1b) > 0:
                st.write("Failed rows for Check 1 (Mode B):")
                failed_view_b = failed_rows_c1b[
                    [
                        key_col_1b,
                        analytics_sso_col,
                        "SSO ID",
                        "Created Date",
                        "Approved Date",
                        "Check1_Result",
                    ]
                ].copy()
                failed_view_b.columns = make_unique(failed_view_b.columns)
                st.dataframe(failed_view_b)
            else:
                st.success("There were no issues or failed cases for this check (Mode B).")

            st.write("Full Check 1 (Mode B) results (ALL rows):")
            preview_cols_b = [
                key_col_1b,
                analytics_sso_col,
                "SSO ID",
                "Created Date",
                "Approved Date",
                "Check1_Result",
            ]
            preview_b = df_check1B[preview_cols_b].copy()
            preview_b.columns = make_unique(preview_b.columns)
            st.dataframe(preview_b)

    # =========================================
    #  3. Check 2 – SOD: Inter-check between Requested / Approvals / Granted
    # =========================================
    st.header("Check 2 – SOD: Requested vs Approvals vs Granted")

    with st.expander("Configure Check 2", expanded=True):
        st.markdown("**Approval levels (up to 5, Leave unused ones blank ).**")
        c2_cols = st.columns(5)
        approval_cols = []

        for i, col in enumerate(c2_cols, start=1):
            with col:
                selected = st.selectbox(
                    f"Approval Level {i} column (optional)",
                    options=[""] + list(df_analytics.columns),
                    key=f"appr_level_{i}"
                )
                approval_cols.append(selected if selected != "" else None)

        requested_by_col = st.selectbox(
            "Requested By column (Analytics)",
            options=df_analytics.columns,
            key="requested_by_col"
        )

        # NEW: format of Requested For
        requested_is_pure_id = st.checkbox(
            "Requested By already contains pure ID (no Name(ID) format)",
            value=True,
            key="requested_is_pure_id_c2",
            help="Tick if Requested By is just the ID (e.g. 223055402), not 'Name (223055402)'."
        )

        st.markdown("**Granted By source (choose one option):**")
        c2_g1, c2_g2 = st.columns(2)
        with c2_g1:
            granted_by_col = st.selectbox(
                "Granted By column (optional)",
                options=[""] + list(df_analytics.columns),
                key="granted_by_col"
            )
        with c2_g2:
            granted_by_id_input = st.text_input(
                "OR enter a single Granted By ID (applies to all rows)",
                value="",
                key="granted_by_id_input"
            )

        # Format of approval & granted columns
        values_are_name_id = st.checkbox(
            "Approval/Granted columns use 'Name (ID)' format (e.g. John Smith (12345))",
            value=True,
            key="values_are_name_id_c2",
            help="If ticked, the tool will extract the ID inside brackets and compare only IDs."
        )

        key_col_2 = st.selectbox(
            "Key/ID column (Analytics, for report)",
            options=df_analytics.columns,
            key="key_col_2"
        )

        run_check2 = st.checkbox("Run Check 2 (for ALL rows)", value=True)

    if run_check2:
        df_check2 = df_analytics.copy()

        # ---------- Normalization helper (to ID) ----------
        def normalize_id(val, treat_as_name_id: bool):
            """
            Convert a cell value to a comparable ID.
            - If NaN/blank -> "" (ignored in comparisons).
            - If treat_as_name_id=True -> extract 'ID' from 'Name (ID)'.
            - Else -> just strip & uppercase.
            """
            if pd.isna(val):
                return ""
            text = str(val).strip()
            if text == "":
                return ""
            if treat_as_name_id:
                return extract_id_from_username(text)
            else:
                return text.upper()

        # ---------- Requested By (ID) ----------
        rb_series = df_check2[requested_by_col]
        if isinstance(rb_series, pd.DataFrame):
            rb_series = rb_series.iloc[:, 0]
        df_check2["_requested_id"] = rb_series.apply(
            lambda v: normalize_id(v, treat_as_name_id=not requested_is_pure_id)
        )

        # ---------- Granted By source ----------
        granted_by_col_selected = granted_by_col if granted_by_col != "" else None

        granted_by_id_manual = granted_by_id_input.strip()
        granted_by_id_manual_norm = (
            normalize_id(granted_by_id_manual, treat_as_name_id=False)
            if granted_by_id_manual
            else ""
        )

        if granted_by_col_selected:
            gb_series = df_check2[granted_by_col_selected]
            if isinstance(gb_series, pd.DataFrame):
                gb_series = gb_series.iloc[:, 0]
            df_check2["_granted_id_col"] = gb_series.apply(
                lambda v: normalize_id(v, treat_as_name_id=values_are_name_id)
            )
        else:
            df_check2["_granted_id_col"] = ""

        df_check2["_granted_id"] = df_check2.apply(
            lambda r: granted_by_id_manual_norm if granted_by_id_manual_norm else r["_granted_id_col"],
            axis=1,
        )

        # ---------- Approvals (ID) ----------
        norm_approval_cols = []
        for idx, col_name in enumerate(approval_cols, start=1):
            if col_name is None:
                continue
            series = df_check2[col_name]
            if isinstance(series, pd.DataFrame):
                series = series.iloc[:, 0]
            norm_name = f"_appr_{idx}_id"
            df_check2[norm_name] = series.apply(
                lambda v: normalize_id(v, treat_as_name_id=values_are_name_id)
            )
            norm_approval_cols.append((col_name, norm_name))

        # ---------- Row-wise SoD check (by ID) ----------
        def sod_intercheck(row):
            req_id = row["_requested_id"]
            grn_id = row["_granted_id"]

            conflicts = []

            if req_id and grn_id and req_id == grn_id:
                conflicts.append("Requested ID = Granted ID")

            for orig_col, norm_col in norm_approval_cols:
                appr_id = row[norm_col]
                if not appr_id:
                    continue

                if req_id and appr_id == req_id:
                    conflicts.append(f"Approval ({orig_col}) ID = Requested ID")
                if grn_id and appr_id == grn_id:
                    conflicts.append(f"Approval ({orig_col}) ID = Granted ID")

            if not conflicts:
                return "Passed", ""

            reason = "Failed - SoD conflict: " + "; ".join(conflicts)
            return reason, "; ".join(conflicts)

        results = df_check2.apply(sod_intercheck, axis=1, result_type="expand")
        df_check2["Check2_Result"] = results[0]
        df_check2["Check2_Conflicts"] = results[1]
        df_check2["Check2_Reason"] = df_check2["Check2_Result"]

        cols_for_output = [key_col_2, requested_by_col]
        if granted_by_col_selected:
            cols_for_output.append(granted_by_col_selected)
        for orig_col, _norm_col in norm_approval_cols:
            cols_for_output.append(orig_col)

        cols_for_output.append("_requested_id")
        cols_for_output.append("_granted_id")
        for _, norm_col in norm_approval_cols:
            cols_for_output.append(norm_col)

        cols_for_output += ["Check2_Result", "Check2_Reason", "Check2_Conflicts"]

        all_results.append(df_check2[cols_for_output])

        total_rows_c2 = len(df_check2)
        failed_rows_c2 = df_check2[df_check2["Check2_Result"].str.startswith("Failed")]
        passed_rows_c2 = total_rows_c2 - len(failed_rows_c2)

        st.subheader("Check 2 Summary")
        st.write(f"Total rows evaluated: {total_rows_c2}")
        st.write(f"Passed: {passed_rows_c2}")
        st.write(f"Failed: {len(failed_rows_c2)}")

        if len(failed_rows_c2) > 0:
            st.write("Failed rows for Check 2:")
            failed_view2 = failed_rows_c2[cols_for_output].copy()
            failed_view2.columns = make_unique(failed_view2.columns)
            st.dataframe(failed_view2)
        else:
            st.success("There were no issues or failed cases for this check.")

        st.write("Full Check 2 results (ALL rows):")
        preview2 = df_check2[cols_for_output].copy()
        preview2.columns = make_unique(preview2.columns)
        st.dataframe(preview2)

    # =========================================
    #  4. Check 3 – Role-by-role comparison (multiple roles per user, ALL rows)
    # =========================================
    st.header("Check 3 – Role Requested = Role Provisioned")

    with st.expander("Configure Check 3 (User & Role Comparison)", expanded=True):
        c3_1, c3_2 = st.columns(2)
        c3_3, c3_4 = st.columns(2)

        with c3_1:
            sbl_userid_col = st.selectbox(
                "User ID column in Dump File",
                options=df_sbl.columns.tolist(),
                key="sbl_userid_col"
            )

        with c3_2:
            sbl_role_col = st.selectbox(
                "Role column in Dump File",
                options=df_sbl.columns.tolist(),
                key="sbl_role_col"
            )

        with c3_3:
            analytics_username_col = st.selectbox(
                "User identifier column in Saviynt  (either Name(ID) or pure ID)",
                options=df_analytics.columns.tolist(),
                key="analytics_username_col"
            )

        with c3_4:
            analytics_role_col = st.selectbox(
                "Role column in Saviynt",
                options=df_analytics.columns.tolist(),
                key="analytics_role_col"
            )

        key_col_3 = st.selectbox(
            "Key/ID column for reporting (from Saviynt Summary)",
            options=df_analytics.columns.tolist(),
            key="key_col_3"
        )

        username_is_pure_id = st.checkbox(
            "Saviynt user column already contains pure ID (no brackets)",
            value=True,
            help="Tick this if the selected Analytics user column is just the ID (e.g., 223055402), "
                 "not 'Name (223055402)'."
        )

        run_check3 = st.checkbox("Run Check 3 (uses ALL rows in both files)", value=True)

    if run_check3:
        sbl_id_role = df_sbl[[sbl_userid_col, sbl_role_col]].copy()

        col_user = sbl_id_role[sbl_userid_col]
        if isinstance(col_user, pd.DataFrame):
            col_user = col_user.iloc[:, 0]

        col_role = sbl_id_role[sbl_role_col]
        if isinstance(col_role, pd.DataFrame):
            col_role = col_role.iloc[:, 0]

        sbl_id_role["_user_id"] = col_user.astype(str).str.strip().str.upper()
        sbl_id_role["_role_norm"] = col_role.astype(str).str.strip().str.upper()

        sbl_map = (
            sbl_id_role.groupby("_user_id")["_role_norm"]
            .apply(lambda x: set(r for r in x if r))
            .to_dict()
        )

        df_check3 = df_analytics[[analytics_username_col, analytics_role_col, key_col_3]].copy()

        user_col_series = df_check3[analytics_username_col]
        if isinstance(user_col_series, pd.DataFrame):
            user_col_series = user_col_series.iloc[:, 0]

        role_col_series = df_check3[analytics_role_col]
        if isinstance(role_col_series, pd.DataFrame):
            role_col_series = role_col_series.iloc[:, 0]

        if username_is_pure_id:
            df_check3["_user_id"] = user_col_series.astype(str).str.strip().str.upper()
        else:
            df_check3["_user_id"] = user_col_series.apply(extract_id_from_username)

        df_check3["_role_norm"] = role_col_series.astype(str).str.strip().str.upper()

        def compare_row(row):
            uid = row["_user_id"]
            role = row["_role_norm"]

            if not uid:
                return "Failed - No ID extracted from username"

            roles_in_sbl = sbl_map.get(uid, None)
            if roles_in_sbl is None or len(roles_in_sbl) == 0:
                return "Failed - User ID not found in Application File"

            if role in roles_in_sbl:
                return "Passed"
            else:
                return "Failed - Role not found for this ID in Application File"

        df_check3["Check3_Result"] = df_check3.apply(compare_row, axis=1)
        df_check3["Check3_Reason"] = df_check3["Check3_Result"]

        all_results.append(
            df_check3[[key_col_3, "_user_id", analytics_role_col, "Check3_Result", "Check3_Reason"]]
        )

        total_rows_c3 = len(df_check3)
        failed_rows_c3 = df_check3[df_check3["Check3_Result"].str.startswith("Failed")]
        passed_rows_c3 = df_check3[df_check3["Check3_Result"] == "Passed"]

        st.subheader("Check 3 Summary")
        st.write(f"Total rows evaluated (Analytics): {total_rows_c3}")
        st.write(f"Passed: {len(passed_rows_c3)}")
        st.write(f"Failed: {len(failed_rows_c3)}")

        st.write("Failed rows for Check 3 (ALL):")
        if not failed_rows_c3.empty:
            failed_view3 = failed_rows_c3[
                [key_col_3, analytics_username_col, "_user_id", analytics_role_col, "Check3_Result"]
            ].copy()
            failed_view3.columns = make_unique(failed_view3.columns)
            st.dataframe(failed_view3)
        else:
            st.success("There were no failed cases for Check 3.")

        st.write("Passed rows for Check 3 (ALL):")
        if not passed_rows_c3.empty:
            passed_view3 = passed_rows_c3[
                [key_col_3, analytics_username_col, "_user_id", analytics_role_col, "Check3_Result"]
            ].copy()
            passed_view3.columns = make_unique(passed_view3.columns)
            st.dataframe(passed_view3)
        else:
            st.info("There were no passed cases for Check 3.")

        st.write("All rows for Check 3 (ALL):")
        all_view3 = df_check3[
            [key_col_3, analytics_username_col, "_user_id", analytics_role_col, "Check3_Result"]
        ].copy()
        all_view3.columns = make_unique(all_view3.columns)
        st.dataframe(all_view3)

    # =========================================
    #  5. Check 4 – Role & Approver Comparison (Per Approval Level vs Entitlement Owner List)
    # =========================================
    st.header("Check 4 – Role Approval Comparison (Saviynt vs Entitlement Owner List)")

    if df_entitlement is None:
        st.info("Upload an Entitlement Owner List file above to enable Check 4.")
    else:
        with st.expander("Configure Check 4", expanded=True):
            st.markdown("**Saviynt (Analytics) side — each approval level is checked separately**")
            c4_0, c4_1 = st.columns(2)

            with c4_0:
                key_col_4 = st.selectbox(
                    "Key/ID column (Saviynt, for report)",
                    options=df_analytics.columns.tolist(),
                    key="key_col_4"
                )

            with c4_1:
                analytics_role_col_c4 = st.selectbox(
                    "Role column (Saviynt)",
                    options=df_analytics.columns.tolist(),
                    key="analytics_role_col_c4"
                )

            analytics_approver_cols_c4 = st.multiselect(
                "Approval Level column(s) in Saviynt — select ALL levels "
                "(e.g., Approval Level 1, Approval Level 2, Approval Level 3...). "
                "Each one will be validated independently.",
                options=df_analytics.columns.tolist(),
                key="analytics_approver_cols_c4"
            )

            analytics_name_id_c4 = st.checkbox(
                "Saviynt approver values use 'Name (ID)' format",
                value=True,
                key="analytics_name_id_c4",
                help="Tick if values look like 'John Smith (12345)' instead of plain names/IDs."
            )

            st.markdown("---")
            st.markdown("**Entitlement Owner List side — used as reference/lookup table (by Role)**")
            c4_3, c4_4 = st.columns(2)

            with c4_3:
                entitlement_role_col = st.selectbox(
                    "Role column (Entitlement Owner List)",
                    options=df_entitlement.columns.tolist(),
                    key="entitlement_role_col"
                )

            with c4_4:
                entitlement_approver_cols = st.multiselect(
                    "Approver / Owner column(s) (Entitlement Owner List)",
                    options=df_entitlement.columns.tolist(),
                    key="entitlement_approver_cols"
                )

            entitlement_name_id_c4 = st.checkbox(
                "Entitlement Owner List values use 'Name (ID)' format",
                value=False,
                key="entitlement_name_id_c4"
            )

            st.markdown("---")
            st.markdown("**Overall row verdict rule**")
            overall_rule_c4 = st.radio(
                "How should the OVERALL row Status be decided from the individual approval levels?",
                options=[
                    "ALL non-blank levels must be valid owners (strict)",
                    "AT LEAST ONE level must be a valid owner (lenient)"
                ],
                index=0,
                key="overall_rule_c4"
            )

            st.markdown("---")
            st.markdown("**Display options**")
            c4_disp1, c4_disp2 = st.columns(2)
            with c4_disp1:
                hide_blank_roles = st.checkbox(
                    "Hide rows with no Role value (still counted in summary)",
                    value=True,
                    key="hide_blank_roles_c4"
                )
            with c4_disp2:
                status_filter_c4 = st.selectbox(
                    "Show rows",
                    options=["All", "Passed only", "Failed only", "Error only", "NA only"],
                    index=0,
                    key="status_filter_c4"
                )

            run_check4 = st.checkbox("Run Check 4 (evaluates ALL rows, ALL approval levels)", value=True, key="run_check4")

        if run_check4:
            if not analytics_approver_cols_c4 or not entitlement_approver_cols:
                st.warning("Please select at least one Approval Level column and one Entitlement Owner column.")
            else:
                # ---------- Build Entitlement Owner reference dict: {role: set(approvers)} ----------
                entitlement_role_dict = build_role_approver_dict(
                    df_entitlement,
                    entitlement_role_col,
                    entitlement_approver_cols,
                    treat_as_name_id=entitlement_name_id_c4
                )

                strict_mode = overall_rule_c4.startswith("ALL")

                # Fixed, guaranteed-unique names for reference/summary columns
                ENTITLEMENT_COL_NAME = "Entitlement Owner Approver(s) [Reference]"

                # ---------- Row-level evaluation: check EACH approval level independently ----------
                def evaluate_check4_row(row):
                    role_val = row[analytics_role_col_c4]
                    role_norm = "" if pd.isna(role_val) or str(role_val).strip() == "" else str(role_val).strip().upper()

                    result = {"Role": role_val if role_norm else "(No Role)"}

                    # Build unique column names per level using a "Saviynt - " prefix,
                    # so they can never collide with the fixed ENTITLEMENT_COL_NAME
                    # or with each other.
                    def approver_col_name(col):
                        return f"Saviynt - {col} (Approver ID)"

                    def valid_col_name(col):
                        return f"Saviynt - {col} (Valid?)"

                    # -------- Case: No role at all -> Status = NA --------
                    if not role_norm:
                        for col in analytics_approver_cols_c4:
                            result[approver_col_name(col)] = "-"
                            result[valid_col_name(col)] = "-"
                        result[ENTITLEMENT_COL_NAME] = "-"
                        result["Status"] = "NA"
                        result["Details"] = "No role value found in this row"
                        return pd.Series(result)

                    entitlement_approvers = entitlement_role_dict.get(role_norm, set())
                    role_exists_in_entitlement = role_norm in entitlement_role_dict

                    # If the role exists in the Entitlement Owner List but has NO
                    # approvers listed against it, there is nothing to validate against —
                    # this is now treated as an "Error" state (data issue), not Pass/Fail.
                    no_owners_defined = role_exists_in_entitlement and len(entitlement_approvers) == 0

                    level_statuses = []       # True/False per non-blank level
                    failed_levels = []        # names of levels that failed
                    passed_levels = []        # names of levels that passed

                    for col in analytics_approver_cols_c4:
                        level_approvers = parse_multi_values(row[col], treat_as_name_id=analytics_name_id_c4)

                        a_col = approver_col_name(col)
                        v_col = valid_col_name(col)

                        if not level_approvers:
                            # blank level -> not counted towards pass/fail
                            result[a_col] = "-"
                            result[v_col] = "-"
                            continue

                        result[a_col] = ", ".join(sorted(level_approvers))

                        if not role_exists_in_entitlement:
                            result[v_col] = "❌ Role not found"
                            level_statuses.append(False)
                            failed_levels.append(col)
                            continue

                        if no_owners_defined:
                            # Nothing to validate against -> mark as N/A at level, contributes to Error status
                            result[v_col] = "⚠️ No owners defined"
                            continue

                        is_valid = level_approvers.issubset(entitlement_approvers) and len(level_approvers) > 0
                        # "valid" here means every approver listed in this level is a recognized entitlement owner
                        if is_valid:
                            result[v_col] = "✅ Valid"
                            level_statuses.append(True)
                            passed_levels.append(col)
                        else:
                            invalid_ones = level_approvers - entitlement_approvers
                            result[v_col] = f"❌ Invalid ({', '.join(sorted(invalid_ones))})"
                            level_statuses.append(False)
                            failed_levels.append(col)

                    result[ENTITLEMENT_COL_NAME] = (
                        ", ".join(sorted(entitlement_approvers)) if entitlement_approvers else "-"
                    )

                    # ---------- Decide overall status ----------
                    if not role_exists_in_entitlement:
                        result["Status"] = "Failed"
                        result["Details"] = "Role not found in Entitlement Owner List"
                    elif no_owners_defined:
                        # Data issue: role exists but has no approvers listed against it -> Error
                        result["Status"] = "Error"
                        result["Details"] = "Role found, but no approvers listed in Entitlement Owner List (auto-passed)"
                    elif not level_statuses:
                        result["Status"] = "Failed"
                        result["Details"] = "No approver value found in any approval level"
                    else:
                        if strict_mode:
                            overall_pass = all(level_statuses)
                        else:
                            overall_pass = any(level_statuses)

                        if overall_pass:
                            result["Status"] = "Passed"
                            result["Details"] = f"Valid level(s): {', '.join(passed_levels)}"
                        else:
                            result["Status"] = "Failed"
                            if strict_mode:
                                result["Details"] = f"Invalid approver(s) at level(s): {', '.join(failed_levels)}"
                            else:
                                result["Details"] = "None of the approval levels matched a valid entitlement owner"

                    return pd.Series(result)

                check4_results = df_analytics.apply(evaluate_check4_row, axis=1)
                df_check4 = pd.concat(
                    [df_analytics[[key_col_4]].reset_index(drop=True), check4_results.reset_index(drop=True)],
                    axis=1
                )
                df_check4.rename(columns={key_col_4: "Request ID"}, inplace=True)

                # ---------- Safety net: guarantee unique column names no matter what ----------
                if df_check4.columns.duplicated().any():
                    df_check4.columns = make_unique(df_check4.columns.tolist())

                # keep for combined download (full, unfiltered) -- includes Valid? columns for audit trail
                all_results.append(df_check4.copy())

                # ---------- Summary ----------
                total_rows_c4 = len(df_check4)
                passed_rows_c4 = df_check4[df_check4["Status"] == "Passed"]
                failed_rows_c4 = df_check4[df_check4["Status"] == "Failed"]
                error_rows_c4 = df_check4[df_check4["Status"] == "Error"]
                na_rows_c4 = df_check4[df_check4["Status"] == "NA"]

                st.subheader("Check 4 Summary")
                m1, m2, m3, m4, m5 = st.columns(5)
                m1.metric("Total Rows", total_rows_c4)
                m2.metric("✅ Passed", len(passed_rows_c4))
                m3.metric("❌ Failed", len(failed_rows_c4))
                m4.metric("⚠️ Error", len(error_rows_c4))
                m5.metric("⬜ NA", len(na_rows_c4))

                # ---------- Apply display filters ----------
                display_df = df_check4.copy()

                if hide_blank_roles:
                    display_df = display_df[display_df["Role"] != "(No Role)"]

                if status_filter_c4 == "Passed only":
                    display_df = display_df[display_df["Status"] == "Passed"]
                elif status_filter_c4 == "Failed only":
                    display_df = display_df[display_df["Status"] == "Failed"]
                elif status_filter_c4 == "Error only":
                    display_df = display_df[display_df["Status"] == "Error"]
                elif status_filter_c4 == "NA only":
                    display_df = display_df[display_df["Status"] == "NA"]

                # ---------- Column order: Request ID, Role, [Approver ID cols only], Entitlement Owners, Status, Details ----------
                # NOTE: "(Valid?)" columns are intentionally EXCLUDED from the on-screen table
                # (they remain in the full df_check4 / CSV download for audit purposes).
                level_cols_ordered = []
                for col in analytics_approver_cols_c4:
                    level_cols_ordered.append(f"Saviynt - {col} (Approver ID)")

                final_col_order = (
                    ["Request ID", "Role"]
                    + level_cols_ordered
                    + [ENTITLEMENT_COL_NAME, "Status", "Details"]
                )
                final_col_order = [c for c in final_col_order if c in display_df.columns]
                display_df = display_df[final_col_order]

                # ---------- Color-coded Status (version-safe styling) ----------
                def highlight_status(val):
                    if val == "Passed":
                        return "background-color: #14532d; color: #dcfce7; font-weight: 600;"
                    elif val == "Failed":
                        return "background-color: #7f1d1d; color: #fee2e2; font-weight: 600;"
                    elif val == "Error":
                        return "background-color: #78350f; color: #fef3c7; font-weight: 600;"
                    elif val == "NA":
                        return "background-color: #374151; color: #d1d5db; font-weight: 600;"
                    return ""

                def highlight_valid_cols(val):
                    if isinstance(val, str):
                        if val.startswith("✅"):
                            return "color: #86efac;"
                        elif val.startswith("❌"):
                            return "color: #fca5a5;"
                        elif val.startswith("⚠️"):
                            return "color: #fde68a;"
                    return ""

                st.write(f"Showing {len(display_df)} of {total_rows_c4} rows")

                # No "(Valid?)" columns remain in display_df, so this will just be an empty list
                valid_cols_to_style = [c for c in display_df.columns if c.endswith("(Valid?)")]

                try:
                    if display_df.columns.duplicated().any():
                        raise ValueError("Duplicate columns detected in display table.")

                    styled_df = style_dataframe_safe(
                        display_df,
                        status_col="Status",
                        valid_cols=valid_cols_to_style,
                        status_style_fn=highlight_status,
                        valid_style_fn=highlight_valid_cols
                    )
                    st.dataframe(styled_df, use_container_width=True, hide_index=True)
                except Exception as style_err:
                    # Fallback: if styling fails for any reason, show plain dataframe instead of crashing
                    st.warning(f"Could not apply cell styling (showing plain table instead). Details: {style_err}")
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                # ---------- Download this check's results separately (FULL data incl. Valid? columns) ----------
                csv_c4 = df_check4.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "📥 Download Check 4 Results (CSV) — includes full audit detail",
                    data=csv_c4,
                    file_name="check4_role_approval_results.csv",
                    mime="text/csv",
                    key="download_check4"
                )

    # =========================================
    #  6. Combined Download (all checks)
    # =========================================
    st.header("Download Combined Results")

    if all_results:
        try:
            combined_df = pd.concat(all_results, ignore_index=True, sort=False)

            st.write(f"Combined report has {len(combined_df)} rows across all checks run.")

            csv_data = combined_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "📥 Download Combined Results (CSV)",
                data=csv_data,
                file_name="sox_access_comparison_results.csv",
                mime="text/csv"
            )
        except Exception as e:
            st.warning(f"Could not combine all results into a single file: {e}")
    else:
        st.info("Run at least one check to enable combined download.")



# ============================================================================
# SECTION 7: CM AUTOMATION TOOL (Change Management Automation)
# ============================================================================


def render_cm_automation_tool():


    TICKET_TYPE_LABELS = {
        "Pulse Tickets": {"Yes": "Normal Ticket", "No": "GXP Standard Ticket"},
        "Other Pulse Tickets": {"Yes": "Standard Ticket", "No": "GXP Normal Ticket"},
    }

    LOG_FILE_PATH = os.path.join(os.getcwd(), "automation_log.txt")

    defaults = {
        "cm_step": 0,
        "cm_sso_id": "",
        "cm_password": "",
        "cm_platform": "Pulse Tickets",
        "cm_ticket_type_key": "Yes",
        "cm_tickets": [],
        "cm_logs": [],
        "cm_download_dir": None,
        "cm_headless": True,
        "cm_excel_bytes": None,
        "cm_excel_filename": "change_management_report.xlsx",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    _log_placeholder = None
    _image_placeholder = None


    def render_process_log(placeholder, logs):
        """Render the process log inside a scrollable, purple-outlined box (display-only helper)."""
        if placeholder is None:
            return
        log_text = "\n".join(logs[-400:])
        escaped = (
            log_text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        placeholder.markdown(
            f'''
            <div style="
                background-color:#FFFFFF;
                border:1.5px solid #5B2A86;
                outline:1.5px solid #5B2A86;
                outline-offset:-1.5px;
                border-radius:8px;
                padding:10px 12px;
                height:400px;
                overflow-y:auto;
                font-family:monospace;
                font-size:13px;
                color:#2B2B2B;
                white-space:pre-wrap;
                word-break:break-word;
            ">{escaped}</div>
            ''',
            unsafe_allow_html=True,
        )


    def log(msg: str):
        timestamped = f"[{now_ist().strftime('%H:%M:%S')}] {msg}"
        st.session_state.cm_logs.append(msg)
        print(msg)
        try:
            with open(LOG_FILE_PATH, "a", encoding="utf-8") as f:
                f.write(timestamped + "\n")
        except Exception as e:
            print(f"Could not write to log file: {e}")
        if _log_placeholder is not None:
            render_process_log(_log_placeholder, st.session_state.cm_logs)


    def show_screenshot(driver, caption: str = ""):
        if _image_placeholder is None:
            return
        try:
            png_bytes = driver.get_screenshot_as_png()
            _image_placeholder.image(png_bytes, caption=caption, use_container_width=True)
        except Exception as e:
            log(f"  (could not capture screenshot: {e})")


    def poll_until(driver, ticket, description, check_fn, timeout=60, interval=2):
        start = time.time()
        while time.time() - start < timeout:
            show_screenshot(driver, f"[{ticket}] {description}...")
            try:
                result = check_fn()
                if result:
                    return result
            except StaleElementReferenceException:
                pass
            except Exception:
                pass
            time.sleep(interval)
        return None


    # ==========================================================
    # Shadow-DOM-aware element finders
    # ==========================================================
    DEEP_QS_JS = """
    function deepQuerySelectorAll(selector, root) {
        root = root || document;
        let results = [];
        try { results = Array.from(root.querySelectorAll(selector)); } catch (e) {}
        const all = root.querySelectorAll('*');
        for (const el of all) {
            if (el.shadowRoot) {
                results = results.concat(deepQuerySelectorAll(selector, el.shadowRoot));
            }
        }
        return results;
    }
    return deepQuerySelectorAll(arguments[0]);
    """

    DEEP_TEXT_JS = """
    function getDirectText(el) {
        let text = '';
        for (const node of el.childNodes) {
            if (node.nodeType === Node.TEXT_NODE) { text += node.textContent; }
        }
        return text.trim();
    }
    function isVisible(el) {
        return !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
    }
    function deepFindByText(text, root) {
        root = root || document;
        const all = root.querySelectorAll('*');
        for (const el of all) {
            if (el.shadowRoot) {
                const res = deepFindByText(text, el.shadowRoot);
                if (res) return res;
            }
            if (getDirectText(el) === text && isVisible(el)) {
                return el;
            }
        }
        return null;
    }
    return deepFindByText(arguments[0]);
    """

    DEEP_ANY_VISIBLE_JS = """
    function isVisible(el) {
        return !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
    }
    function deepAnyVisible(selector, root) {
        root = root || document;
        let found = [];
        try { found = Array.from(root.querySelectorAll(selector)); } catch (e) {}
        for (const el of found) {
            if (isVisible(el)) return true;
        }
        const all = root.querySelectorAll('*');
        for (const el of all) {
            if (el.shadowRoot) {
                if (deepAnyVisible(selector, el.shadowRoot)) return true;
            }
        }
        return false;
    }
    return deepAnyVisible(arguments[0]);
    """

    DEEP_TEXT_CONTAINS_JS = """
    function getDirectText(el) {
        let text = '';
        for (const node of el.childNodes) {
            if (node.nodeType === Node.TEXT_NODE) { text += node.textContent; }
        }
        return text.trim();
    }
    function isVisible(el) {
        return !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
    }
    function deepFindByTextContains(text, root) {
        root = root || document;
        const all = root.querySelectorAll('*');
        for (const el of all) {
            if (el.shadowRoot) {
                const res = deepFindByTextContains(text, el.shadowRoot);
                if (res) return res;
            }
            const t = getDirectText(el);
            if (t.indexOf(text) !== -1 && isVisible(el)) {
                return el;
            }
        }
        return null;
    }
    return deepFindByTextContains(arguments[0]);
    """


    def deep_find_all(driver, css_selector):
        try:
            return driver.execute_script(DEEP_QS_JS, css_selector) or []
        except Exception:
            return []


    def deep_find(driver, css_selector):
        els = deep_find_all(driver, css_selector)
        for el in els:
            try:
                if el.is_displayed():
                    return el
            except Exception:
                continue
        return els[0] if els else None


    def deep_find_by_text(driver, text):
        try:
            return driver.execute_script(DEEP_TEXT_JS, text)
        except Exception:
            return None


    def deep_find_by_text_contains(driver, text):
        try:
            return driver.execute_script(DEEP_TEXT_CONTAINS_JS, text)
        except Exception:
            return None


    def deep_any_visible(driver, css_selector):
        try:
            return bool(driver.execute_script(DEEP_ANY_VISIBLE_JS, css_selector))
        except Exception:
            return False


    def safe_click(driver, el):
        try:
            el.click()
            return
        except Exception:
            pass
        try:
            driver.execute_script("arguments[0].click();", el)
            return
        except Exception:
            pass
        try:
            ActionChains(driver).move_to_element(el).click().perform()
        except Exception as e:
            raise e


    # ==========================================================
    # Iframe-aware search wrapper (iframes found via shadow-DOM-aware search)
    # ==========================================================
    def find_across_frames(driver, finder_fn, max_depth=3):
        el = finder_fn(driver)
        if el:
            return el

        driver.switch_to.default_content()
        el = finder_fn(driver)
        if el:
            return el

        def search_frames(depth):
            if depth > max_depth:
                return None
            frames = deep_find_all(driver, "iframe")
            for frame in frames:
                try:
                    driver.switch_to.frame(frame)
                except Exception:
                    continue
                found = finder_fn(driver)
                if found:
                    return found
                nested = search_frames(depth + 1)
                if nested:
                    return nested
                driver.switch_to.parent_frame()
            return None

        result = search_frames(1)
        if result:
            return result

        driver.switch_to.default_content()
        return None


    def find_all_across_frames(driver, finder_fn, max_depth=3):
        """Like find_across_frames, but collects ALL matches from every frame
        (top-level + every nested iframe, found via shadow-DOM-aware search).
        NOTE: elements returned here may go stale once the driver switches
        frames again -- do not read attributes off them after calling this."""
        results = []

        def collect():
            try:
                found = finder_fn(driver)
            except Exception:
                found = None
            if found:
                results.extend(found)

        driver.switch_to.default_content()
        collect()

        def search_frames(depth):
            if depth > max_depth:
                return
            frames = deep_find_all(driver, "iframe")
            for frame in frames:
                try:
                    driver.switch_to.frame(frame)
                except Exception:
                    continue
                collect()
                search_frames(depth + 1)
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    driver.switch_to.default_content()

        search_frames(1)
        driver.switch_to.default_content()
        return results


    # ==========================================================
    # Loading-spinner / step-completion helpers
    # ==========================================================
    SPINNER_SELECTORS = [
        "[class*='spinner' i]",
        "[class*='loading' i]",
        "[aria-busy='true']",
        ".now-loading-indicator",
    ]


    def is_loading(driver):
        for sel in SPINNER_SELECTORS:
            if deep_any_visible(driver, sel):
                return True
        return False


    def wait_for_spinner_to_clear(driver, ticket, context_label, timeout=30, interval=1):
        start = time.time()
        while time.time() - start < timeout:
            if not is_loading(driver):
                return True
            show_screenshot(driver, f"[{ticket}] {context_label}: waiting for loading to finish...")
            time.sleep(interval)
        return not is_loading(driver)


    def wait_until_gone(driver, ticket, description, css_selector, timeout=30, interval=1):
        start = time.time()
        while time.time() - start < timeout:
            if not deep_any_visible(driver, css_selector):
                return True
            show_screenshot(driver, f"[{ticket}] {description}: waiting to close...")
            time.sleep(interval)
        return not deep_any_visible(driver, css_selector)


    # ==========================================================
    # Diagnostics
    # ==========================================================
    def debug_dump_page_state(driver, ticket):
        try:
            driver.switch_to.default_content()
            log(f"  [DEBUG] === Diagnostics for ticket {ticket} ===")
            log(f"  [DEBUG] Current URL: {driver.current_url}")
            log(f"  [DEBUG] Page title: {driver.title}")
            log(f"  [DEBUG] Number of window handles: {len(driver.window_handles)}")

            iframes = deep_find_all(driver, "iframe")
            log(f"  [DEBUG] Found {len(iframes)} iframe(s) on the page (incl. shadow DOM):")
            for i, f in enumerate(iframes):
                try:
                    fid = f.get_attribute("id")
                    fsrc = f.get_attribute("src")
                    fname = f.get_attribute("name")
                    log(f"    iframe[{i}] id={fid!r} name={fname!r} src={str(fsrc)[:150]!r}")
                except Exception as e:
                    log(f"    iframe[{i}] (could not read attributes: {e})")

            candidate_selectors = [
                "button.additional-actions-context-menu-button",
                "button[aria-label='additional actions']",
                "button[aria-label*='action' i]",
                "button[title*='action' i]",
                "[aria-haspopup='true']",
                "button.icon-menu",
            ]

            def scan_frame(label):
                for sel in candidate_selectors:
                    els = deep_find_all(driver, sel)
                    if els:
                        details = []
                        for e in els:
                            try:
                                details.append(
                                    f"(visible={e.is_displayed()}, text={e.text!r}, "
                                    f"aria-label={e.get_attribute('aria-label')!r})"
                                )
                            except Exception as ex:
                                details.append(f"(error reading: {ex})")
                        log(f"    [{label}] {sel!r} -> {len(els)} match(es): {details}")
                    else:
                        log(f"    [{label}] {sel!r} -> 0 matches")

            scan_frame("top-level")
            for i, f in enumerate(iframes):
                try:
                    driver.switch_to.frame(f)
                    scan_frame(f"iframe[{i}]")
                    driver.switch_to.default_content()
                except Exception as e:
                    log(f"    [iframe[{i}]] could not switch into frame: {e}")
                    driver.switch_to.default_content()

            log(f"  [DEBUG] === End diagnostics for {ticket} ===")

        except Exception as e:
            log(f"  [DEBUG] debug_dump_page_state itself failed: {e}")
        finally:
            driver.switch_to.default_content()


    # ==========================================================
    # Binary detection (HARDENED: known paths -> PATH -> filesystem-wide search)
    # ==========================================================
    def _filesystem_search(patterns, roots=("/usr", "/opt", "/snap", "/root", "/home")):
        """Last-resort recursive search for a binary matching any of `patterns`."""
        for root in roots:
            if not os.path.isdir(root):
                continue
            for pattern in patterns:
                try:
                    matches = glob.glob(os.path.join(root, "**", pattern), recursive=True)
                except Exception:
                    matches = []
                for m in matches:
                    if os.path.isfile(m) and os.access(m, os.X_OK) or os.path.isfile(m):
                        return m
        return None


    def find_browser_binary():
        candidates = [
            "/usr/bin/chromium",
            "/usr/bin/chromium-browser",
            "/usr/bin/google-chrome",
            "/usr/bin/google-chrome-stable",
            "/usr/lib/chromium/chromium",
            "/usr/lib/chromium-browser/chromium-browser",
            "/snap/bin/chromium",
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        for name in ["chromium", "chromium-browser", "google-chrome", "google-chrome-stable"]:
            found = shutil.which(name)
            if found:
                return found
        found = _filesystem_search(["chromium", "chromium-browser", "google-chrome*"])
        return found


    def find_chromedriver_binary():
        candidates = [
            "/usr/bin/chromedriver",
            "/usr/lib/chromium/chromedriver",
            "/usr/lib/chromium-browser/chromedriver",
            "/snap/bin/chromedriver",
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        found = shutil.which("chromedriver")
        if found:
            return found
        found = _filesystem_search(["chromedriver"])
        if found:
            try:
                os.chmod(found, 0o755)
            except Exception:
                pass
        return found


    def run_environment_diagnostics():
        log("=== Environment diagnostics ===")
        try:
            which_chromium = shutil.which("chromium") or shutil.which("chromium-browser")
            which_chromedriver = shutil.which("chromedriver")
            log(f"  which chromium: {which_chromium}")
            log(f"  which chromedriver: {which_chromedriver}")
        except Exception as e:
            log(f"  which check failed: {e}")

        try:
            result = subprocess.run(
                ["dpkg", "-l"], capture_output=True, text=True, timeout=10
            )
            lines = [l for l in result.stdout.splitlines() if "chrom" in l.lower()]
            if lines:
                log("  dpkg packages matching 'chrom':")
                for l in lines:
                    log(f"    {l}")
            else:
                log("  dpkg: no chromium/chromedriver packages found "
                    "(packages.txt may be missing or app wasn't rebooted after adding it)")
        except Exception as e:
            log(f"  dpkg check failed or unavailable: {e}")

        log("=== End environment diagnostics ===")


    # ==========================================================
    # Selenium driver setup
    # ==========================================================
    def build_driver(download_dir: str, headless: bool = True):
        run_environment_diagnostics()

        options = Options()
        if headless:
            options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")

        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)

        chromium_path = find_browser_binary()
        chromedriver_path = find_chromedriver_binary()

        log(f"Detected chromium binary: {chromium_path}")
        log(f"Detected chromedriver binary: {chromedriver_path}")

        if chromium_path:
            options.binary_location = chromium_path

        driver = None

        if chromedriver_path:
            try:
                service = Service(chromedriver_path)
                driver = webdriver.Chrome(service=service, options=options)
                log(f"Using detected chromedriver at {chromedriver_path}")
            except Exception as e:
                log(f"Detected chromedriver failed to launch: {e}")
        else:
            log("No chromedriver binary found via known paths, system PATH, or filesystem search. "
                "This almost certainly means packages.txt is missing 'chromium-driver' "
                "or the app has not been rebooted since adding it.")

        if driver is None:
            try:
                driver = webdriver.Chrome(options=options)
                log("Using Selenium Manager auto-resolved driver")
            except Exception as e:
                log(f"Selenium Manager failed: {e}")

        if driver is None:
            try:
                from webdriver_manager.chrome import ChromeDriverManager
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=options)
                log("Using webdriver-manager driver")
            except Exception as e:
                log(f"webdriver-manager also failed: {e}")
                raise RuntimeError(
                    "Could not start Chrome/Chromium via any method. "
                    "Check that packages.txt contains 'chromium' and 'chromium-driver', "
                    "and that you have rebooted the app (Manage app -> Reboot) after adding it."
                )

        try:
            driver.execute_cdp_cmd("Page.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": download_dir
            })
            log(f"Download behavior explicitly set via CDP to: {download_dir}")
        except Exception as e:
            log(f"Warning: could not set CDP download behavior: {e}")

        return driver


    # ==========================================================
    # Login — fully automated username + password, then wait for MFA push
    # ==========================================================
    def login(driver, sso_id: str, password: str):
        login_url = "https://pulse.service-now.com/now/nav/ui/home"
        expected_domain = "pulse.service-now.com"

        log("Opening login page...")
        driver.get(login_url)
        show_screenshot(driver, "Login page loaded")

        try:
            username_field = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='text']"))
            )
            username_field.send_keys(sso_id)
            show_screenshot(driver, "Username entered")
            log("Entered SSO ID, clicking Next...")

            next_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Next')]"))
            )
            next_btn.click()
        except TimeoutException:
            show_screenshot(driver, "FAILED at username step")
            raise TimeoutException("Could not find username field or Next button on login page")

        log("Waiting 2 seconds before entering password...")
        time.sleep(2)
        show_screenshot(driver, "Waiting before password entry")

        try:
            password_field = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='password']"))
            )
            password_field.send_keys(password)
            show_screenshot(driver, "Password entered")
            log("Entered password, submitting...")

            login_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Log In & Remember Me')]"))
            )
            login_btn.click()
            log("Clicked 'Log In & Remember Me'")
        except TimeoutException:
            show_screenshot(driver, "FAILED at password step")
            raise TimeoutException("Could not find password field or Login button")

        log("Password submitted. A push notification should now be sent to your "
            "authenticator app. Please approve it on your phone.")
        log("Waiting for you to approve the push notification... (up to 5 minutes)")

        start = time.time()
        timeout = 300
        while time.time() - start < timeout:
            if expected_domain in driver.current_url and "login" not in driver.current_url.lower():
                break
            show_screenshot(driver, "Waiting for MFA push approval on your phone...")
            time.sleep(3)
        else:
            raise TimeoutException(
                "Login was not completed within 5 minutes. "
                "Did you approve the push notification on your phone?"
            )

        log("Login successful! Resuming automation...")
        show_screenshot(driver, "Logged in successfully")


    # ==========================================================
    # Step 1: search box (shadow-DOM aware)
    # ==========================================================
    def open_search_and_type(driver, ticket):
        def find_input():
            return deep_find(driver, "#sncwsgs-typeahead-input")

        search_box = poll_until(driver, ticket, "Step 1a: finding search input", find_input, timeout=15, interval=1)

        if search_box is None:
            def find_trigger():
                for sel in ["input[placeholder='Search']", "[aria-label='Search']",
                            "button[aria-label*='search' i]", ".search-box", ".global-search"]:
                    el = deep_find(driver, sel)
                    if el:
                        return el
                return None

            trigger = poll_until(driver, ticket, "Step 1a: finding search trigger", find_trigger, timeout=15, interval=1)
            if trigger is None:
                raise TimeoutException("Could not find search box/icon (even searching shadow DOM)")
            safe_click(driver, trigger)
            show_screenshot(driver, f"[{ticket}] Step 1a: clicked search trigger")

            search_box = poll_until(driver, ticket, "Step 1b: waiting for expanded search input", find_input, timeout=15, interval=1)
            if search_box is None:
                raise TimeoutException("Expanded search input never appeared after clicking search")

        safe_click(driver, search_box)
        search_box.clear()
        search_box.send_keys(ticket)
        search_box.send_keys(Keys.ENTER)
        show_screenshot(driver, f"[{ticket}] Step 1: searched, waiting for results")
        return search_box


    # ==========================================================
    # Per-ticket folder helpers
    # ==========================================================
    def sanitize_ticket_name(ticket: str) -> str:
        cleaned = "".join(c for c in ticket.strip() if c.isalnum() or c in ("-", "_"))
        return cleaned or "ticket"


    def get_ticket_folder(download_dir: str, ticket: str) -> str:
        """Every ticket gets its own dedicated folder: <download_dir>/<TICKET>/
        The ticket's PDF goes directly inside it, and its attachments go inside
        a further 'attachments' subfolder."""
        folder = os.path.join(download_dir, sanitize_ticket_name(ticket))
        os.makedirs(folder, exist_ok=True)
        return folder


    def get_ticket_attachments_folder(download_dir: str, ticket: str) -> str:
        folder = os.path.join(get_ticket_folder(download_dir, ticket), "attachments")
        os.makedirs(folder, exist_ok=True)
        return folder


    def cleanup_empty_attachments_folder(attachments_dir: str):
        """If a ticket genuinely had zero attachments, the 'attachments'
        subfolder was still created up-front (needed as a destination in case
        there WERE any). Remove it afterward if it ended up empty, so tickets
        with no attachments don't leave a clutter empty folder behind in the
        final ZIP/download listing."""
        try:
            if os.path.isdir(attachments_dir) and not os.listdir(attachments_dir):
                os.rmdir(attachments_dir)
        except Exception:
            pass


    def build_attachment_filename(ticket, original_name):
        """Prefix the ticket number onto the attachment's original filename,
        unless it's already present (ServiceNow sometimes already includes it)."""
        original_name = (original_name or "attachment").strip()
        if original_name.upper().startswith(ticket.strip().upper()):
            return original_name
        return f"{ticket.strip()}_{original_name}"


    # ==========================================================
    # Attachment download helpers
    # ==========================================================
    ATTACHMENT_LINK_SELECTOR_PRIMARY = "a[aria-label^='Download ' i]"
    ATTACHMENT_LINK_SELECTOR_FALLBACK = "a[href*='sys_attachment.do']"


    def _attachment_link_name(el):
        """Derive a filename for an attachment <a> element. MUST be called
        immediately after locating the element, in the same frame context it
        was found in."""
        name = None
        try:
            aria = el.get_attribute("aria-label") or ""
        except Exception:
            aria = ""
        if aria.lower().startswith("download "):
            name = aria[len("Download "):].strip()
        if not name:
            try:
                text = (el.text or "").strip()
            except Exception:
                text = ""
            if text:
                name = text
        if not name:
            try:
                href = el.get_attribute("href") or ""
            except Exception:
                href = ""
            if href:
                name = href.split("/")[-1].split("?")[0]
        return name or None


    def _is_view_link(el):
        """Filters out the separate 'View <filename>' links that sit next to
        each 'Download <filename>' link -- we only want to click Download."""
        try:
            aria = (el.get_attribute("aria-label") or "").lower()
            if aria.startswith("view "):
                return True
        except Exception:
            pass
        try:
            cls = (el.get_attribute("class") or "").lower()
            if "view_" in cls or "view-" in cls:
                return True
        except Exception:
            pass
        return False


    def find_next_unclicked_attachment(driver, downloaded_names):
        """Finds the first attachment Download link not yet in `downloaded_names`,
        returning it WITHOUT switching frames afterward, so it's still valid for
        an immediate click()."""

        def finder(d):
            els = deep_find_all(d, ATTACHMENT_LINK_SELECTOR_PRIMARY)
            if not els:
                els = [e for e in deep_find_all(d, ATTACHMENT_LINK_SELECTOR_FALLBACK)
                       if not _is_view_link(e)]
            for el in els:
                name = _attachment_link_name(el)
                if name and name not in downloaded_names:
                    el._attachment_name_cache = name
                    return el
            return None

        return find_across_frames(driver, finder)


    def open_manage_attachments(driver, ticket):
        """Some tickets show every attachment inline in the header already
        (no popup needed); others truncate the list and require clicking
        'Manage Attachments (N)' to reveal the rest. Returns True if any
        Download link is present anywhere on the page afterward, False if the
        ticket has zero attachments (this is a NORMAL, expected outcome, not
        an error)."""

        def check_any_link(d):
            el = deep_find(d, ATTACHMENT_LINK_SELECTOR_PRIMARY)
            if el:
                return el
            for c in deep_find_all(d, ATTACHMENT_LINK_SELECTOR_FALLBACK):
                if not _is_view_link(c):
                    return c
            return None

        already_visible = find_across_frames(driver, check_any_link)
        if already_visible:
            log(f"  Step 2c: attachment link(s) already visible inline for {ticket}")
            driver.switch_to.default_content()
            return True

        def find_trigger(d):
            el = deep_find(d, "a#header_attachment_list_label")
            if el:
                return el
            el = deep_find_by_text_contains(d, "Manage Attachments")
            if el:
                return el
            return deep_find_by_text_contains(d, "Attachments (")

        trigger = poll_until(
            driver, ticket, "Step 2c: locating 'Manage Attachments' link",
            lambda: find_across_frames(driver, find_trigger),
            timeout=15, interval=1
        )
        if trigger is None:
            log(f"  Step 2c: 'Manage Attachments' link NOT FOUND for {ticket} "
                f"(this ticket has 0 attachments -- normal, not an error)")
            driver.switch_to.default_content()
            return False

        log(f"  Step 2c: found 'Manage Attachments' trigger for {ticket}, clicking it")
        try:
            safe_click(driver, trigger)
        except Exception as e:
            log(f"  Step 2c: could not click 'Manage Attachments' trigger for {ticket}: {e}")
            driver.switch_to.default_content()
            return False
        show_screenshot(driver, f"[{ticket}] Step 2c: opened Manage Attachments dialog")

        opened = poll_until(
            driver, ticket, "Step 2c: waiting for Attachments dialog to load",
            lambda: find_across_frames(driver, check_any_link),
            timeout=20, interval=1
        )
        driver.switch_to.default_content()
        if not opened:
            log(f"  Step 2c: Attachments popup opened but no Download links "
                f"rendered within 20s for {ticket} (ticket may genuinely have 0 "
                f"attachments, or the popup listed them differently than expected)")
        return bool(opened)


    def close_manage_attachments_dialog(driver, ticket):
        """Closes the popup via an explicit close/X button only. Deliberately
        never sends ESCAPE -- in ServiceNow's classic UI, ESC can navigate away
        from the ticket entirely."""

        def check_open(d):
            el = deep_find(d, ATTACHMENT_LINK_SELECTOR_PRIMARY)
            if el:
                return el
            return deep_find(d, "div.modal-dialog, div[role='dialog']")

        still_open = find_across_frames(driver, check_open)
        if not still_open:
            driver.switch_to.default_content()
            return

        def find_close_btn(d):
            for sel in ["button[aria-label='Close']", "button[title='Close']",
                        ".close", "[aria-label*='close' i]", "button.btn-close",
                        "img[alt='Close']", "a[title='Close']", ".glide_dialog_close"]:
                el = deep_find(d, sel)
                if el:
                    return el
            return None

        btn = find_across_frames(driver, find_close_btn)
        if btn:
            try:
                safe_click(driver, btn)
                log(f"  Closed Attachments dialog for {ticket}")
            except Exception as e:
                log(f"  Could not close Attachments dialog for {ticket}: {e} (leaving it open)")
        else:
            log(f"  No explicit close button found for Attachments dialog on {ticket}; "
                f"leaving it open (non-blocking) rather than sending ESCAPE.")

        driver.switch_to.default_content()
        time.sleep(1)


    def download_ticket_attachments(driver, ticket, download_dir, attachments_dir):
        """Finds and downloads every attachment for the ticket ONE AT A TIME.
        If the ticket has zero attachments, this cleanly logs that and returns
        -- no error, no retries, no impact on the PDF that was already
        downloaded before this step runs."""

        opened = open_manage_attachments(driver, ticket)
        if not opened:
            log(f"  No attachments found for {ticket} (skipping attachment download)")
            return

        log(f"  Step 2c: attachment link(s) confirmed present for {ticket}, downloading...")
        downloaded_names = set()
        max_iterations = 25
        consecutive_click_failures = 0

        for _ in range(max_iterations):
            target = find_next_unclicked_attachment(driver, downloaded_names)

            if target is None:
                if not downloaded_names:
                    log(f"  Step 2c: WARNING -- no attachment Download links could be "
                        f"matched for {ticket} even though the trigger/popup opened.")
                break

            target_name = getattr(target, "_attachment_name_cache", None) or "attachment"

            log(f"  Step 2c: downloading attachment '{target_name}' for {ticket}")
            before_files = _list_files_only(get_candidate_download_dirs(download_dir))

            try:
                safe_click(driver, target)
                consecutive_click_failures = 0
            except Exception as e:
                log(f"    Could not click attachment link '{target_name}' for {ticket}: {e}")
                downloaded_names.add(target_name)
                consecutive_click_failures += 1
                driver.switch_to.default_content()
                if consecutive_click_failures >= 3:
                    log(f"    Too many consecutive click failures for {ticket}, stopping")
                    break
                continue

            driver.switch_to.default_content()

            show_screenshot(driver, f"[{ticket}] downloading attachment: {target_name}")
            downloaded_file = wait_for_new_download(download_dir, before_files, timeout=90)

            if downloaded_file:
                new_name = build_attachment_filename(ticket, target_name)
                dest_path = os.path.join(attachments_dir, new_name)
                if os.path.exists(dest_path):
                    root_name, ext = os.path.splitext(new_name)
                    dest_path = os.path.join(attachments_dir, f"{root_name}_{int(time.time())}{ext}")
                try:
                    shutil.move(downloaded_file, dest_path)
                    log(f"    Attachment saved as: {os.path.relpath(dest_path, download_dir)}")
                except Exception as e:
                    log(f"    Attachment downloaded but move/rename failed for {ticket}: {e}. "
                        f"File left at: {downloaded_file}")
            else:
                log(f"    Attachment '{target_name}' did not finish downloading within 90s for {ticket}")

            downloaded_names.add(target_name)
            time.sleep(1)
        else:
            log(f"  Reached max attachment-download iterations ({max_iterations}) for {ticket}")

        log(f"  Finished attachment downloads for {ticket}: {len(downloaded_names)} attachment(s) handled")
        close_manage_attachments_dialog(driver, ticket)


    # ==========================================================
    # Full ticket download flow
    # ==========================================================
    def download_ticket_pdf(driver, ticket: str, download_dir: str):
        driver.switch_to.default_content()

        log("  Step 1: clicking search box and typing ticket")
        open_search_and_type(driver, ticket)

        log("  Step 2: waiting for loading spinner to clear, then results")
        wait_for_spinner_to_clear(driver, ticket, "Step 2 (search)", timeout=30)

        def check_results(d):
            if is_loading(d):
                return None
            el = deep_find(d, "ul[aria-labelledby='section-EXACT_MATCH_SECTION'] li[data-testclass='sn-global-search-record']")
            if el:
                return el
            return deep_find(d, "li[data-testclass='sn-global-search-record']")

        result = poll_until(
            driver, ticket, "Step 2: waiting for search results",
            lambda: check_results(driver), timeout=60, interval=2
        )
        if result is None:
            show_screenshot(driver, f"[{ticket}] Step 2: FAILED — no results after 60s")
            raise TimeoutException(f"No search results appeared for ticket {ticket} within 60s")

        log("  Step 2: result found, clicking it")
        safe_click(driver, result)
        show_screenshot(driver, f"[{ticket}] Step 2: opened ticket")

        time.sleep(3)

        log("  Step 2b: confirming ticket detail page loaded")
        wait_for_spinner_to_clear(driver, ticket, "Step 2b (opening ticket)", timeout=30)

        menu_selectors = [
            "button.additional-actions-context-menu-button",
            "button[aria-label='additional actions']",
            "button[aria-label='Additional actions']",
            "button[aria-label*='additional actions' i]",
            "button[title*='additional actions' i]",
        ]

        def check_menu_button(d):
            if is_loading(d):
                return None
            for sel in menu_selectors:
                el = deep_find(d, sel)
                if el:
                    return el
            return None

        menu_btn = poll_until(
            driver, ticket, "Step 2b: confirming ticket page loaded",
            lambda: find_across_frames(driver, check_menu_button),
            timeout=40, interval=2
        )
        if menu_btn is None:
            log(f"  Step 2b: FAILED for {ticket} — running diagnostics...")
            debug_dump_page_state(driver, ticket)
            show_screenshot(driver, f"[{ticket}] Step 2b: FAILED — see automation_log.txt")
            raise TimeoutException(
                f"Ticket detail page never finished loading for {ticket}. "
                f"See {LOG_FILE_PATH} for diagnostics."
            )

        log("  Step 3: opening 'Additional actions' menu")
        fresh_menu_btn = find_across_frames(driver, check_menu_button) or menu_btn
        safe_click(driver, fresh_menu_btn)
        show_screenshot(driver, f"[{ticket}] Step 3: menu opened")

        time.sleep(1)

        log("  Step 4: hovering 'Export'")

        def check_export_item(d):
            return deep_find_by_text(d, "Export")

        export_item = poll_until(
            driver, ticket, "Step 4: waiting for Export menu item",
            lambda: find_across_frames(driver, check_export_item),
            timeout=20, interval=1
        )
        if export_item is None:
            debug_dump_page_state(driver, ticket)
            raise TimeoutException(f"'Export' menu item never appeared for ticket {ticket}")
        ActionChains(driver).move_to_element(export_item).perform()
        show_screenshot(driver, f"[{ticket}] Step 4: hovering Export")

        log("  Step 5: clicking 'PDF' in flyout")

        def check_pdf_item(d):
            return deep_find_by_text(d, "PDF")

        pdf_item = poll_until(
            driver, ticket, "Step 5: waiting for PDF flyout item",
            lambda: find_across_frames(driver, check_pdf_item),
            timeout=20, interval=1
        )
        if pdf_item is None:
            debug_dump_page_state(driver, ticket)
            raise TimeoutException(f"'PDF' flyout item never appeared for ticket {ticket}")
        safe_click(driver, pdf_item)
        show_screenshot(driver, f"[{ticket}] Step 5: clicked PDF")

        log("  Step 5b: confirming 'Export to PDF' dialog fully opened")

        def check_dialog_open(d):
            return deep_find(d, "#ok_button")

        ok_btn = poll_until(
            driver, ticket, "Step 5b: waiting for Export dialog to appear",
            lambda: find_across_frames(driver, check_dialog_open),
            timeout=20, interval=1
        )
        if ok_btn is None:
            raise TimeoutException(f"Export dialog 'ok_button' never appeared for ticket {ticket}")

        log("  Step 6: clicking Export button in dialog")
        safe_click(driver, ok_btn)
        show_screenshot(driver, f"[{ticket}] Step 6: export triggered")

        log("  Step 6b: confirming orientation dialog closed")
        wait_until_gone(driver, ticket, "Step 6b (orientation dialog)", "#ok_button", timeout=15)

        log("  Step 7: waiting for Download button to appear AND become enabled "
            "(PDF generation can take time)")

        _logged_once = {"done": False}

        def check_download_button(d):
            if is_loading(d):
                return None
            el = deep_find(d, "#download_button")
            if el is None:
                return None
            try:
                disabled_attr = el.get_attribute("disabled")
                aria_disabled = el.get_attribute("aria-disabled")
                classes = (el.get_attribute("class") or "").lower()
                if not _logged_once["done"]:
                    log(f"    [debug] download_button found: disabled={disabled_attr!r} "
                        f"aria-disabled={aria_disabled!r} class={classes!r} "
                        f"is_enabled={el.is_enabled()}")
                    _logged_once["done"] = True
                if disabled_attr or aria_disabled == "true" or "disabled" in classes:
                    return None
                if not el.is_enabled():
                    return None
            except Exception:
                return None
            return el

        download_btn = poll_until(
            driver, ticket, "Step 7: waiting for PDF generation to finish (button enabled)",
            lambda: find_across_frames(driver, check_download_button),
            timeout=120, interval=3
        )
        if download_btn is None:
            raise TimeoutException(f"Download button never became enabled for ticket {ticket}")
        show_screenshot(driver, f"[{ticket}] Step 7: export ready (button enabled)")

        time.sleep(1)

        safe_click(driver, download_btn)
        log("  Step 8: clicked Download")
        show_screenshot(driver, f"[{ticket}] Step 8: downloading...")

        log("  Step 8b: waiting for 'Export Complete' dialog to close on its own "
            "(this can take a while depending on file size)")
        dialog_closed = wait_until_gone(
            driver, ticket, "Step 8b (export complete dialog)", "#download_button", timeout=90
        )
        if dialog_closed:
            log(f"  Step 8b: dialog closed normally for {ticket}")
        else:
            log(f"  Step 8b: dialog did NOT close within 90s for {ticket} — "
                f"will check if the file downloaded anyway, and try to force-close the dialog.")
        show_screenshot(driver, f"[{ticket}] Step 8b: done waiting on dialog")

        driver.switch_to.default_content()

        log("  Step 9: downloading all attachments (if any) after PDF export")
        attachments_dir = get_ticket_attachments_folder(download_dir, ticket)
        try:
            download_ticket_attachments(driver, ticket, download_dir, attachments_dir)
        except Exception as ex:
            log(f"  Attachment download step failed for {ticket}: {type(ex).__name__}: {ex}")
            show_screenshot(driver, f"[{ticket}] Attachment download step failed")
        finally:
            # If the ticket had 0 attachments, don't leave a clutter empty
            # "attachments" folder behind in the final output.
            cleanup_empty_attachments_folder(attachments_dir)

        driver.switch_to.default_content()


    # ==========================================================
    # Download detection — checks multiple candidate folders as a safety net
    # ==========================================================
    def get_candidate_download_dirs(download_dir: str):
        candidates = [download_dir]
        home = os.path.expanduser("~")
        default_downloads = os.path.join(home, "Downloads")
        if default_downloads not in candidates:
            candidates.append(default_downloads)
        return candidates


    def _list_files_only(dirs):
        """Returns only FILES (never directories) found directly inside any of
        `dirs`. Critical: per-ticket folders (e.g. <download_dir>/CHG12345/ and
        its attachments/ subfolder) get created via os.makedirs() -- a plain
        glob.glob(dir/"*") would also match those new folders, which could get
        mistaken for the downloaded PDF file itself. Filtering to
        os.path.isfile() only avoids that."""
        result = set()
        for d in dirs:
            for p in glob.glob(os.path.join(d, "*")):
                if os.path.isfile(p):
                    result.add(p)
        return result


    def wait_for_new_download(download_dir: str, before_files: set, timeout: int = 120):
        candidate_dirs = get_candidate_download_dirs(download_dir)
        start = time.time()
        while time.time() - start < timeout:
            current_files = _list_files_only(candidate_dirs)
            new_files = current_files - before_files
            finished = [f for f in new_files if not f.endswith(".crdownload") and not f.endswith(".tmp")]
            still_downloading = any(
                f.endswith(".crdownload") or f.endswith(".tmp") for f in current_files
            )
            if finished and not still_downloading:
                return finished[0]
            time.sleep(1)
        return None


    def rename_downloaded_file(filepath: str, ticket_number: str, ticket_type_label: str, download_dir: str):
        """Moves the exported PDF into the ticket's dedicated folder
        (<download_dir>/<TICKET>/<TICKET>_<label>.pdf), separate from its
        attachments (which live in <download_dir>/<TICKET>/attachments/)."""
        if not filepath or not os.path.exists(filepath):
            return None
        safe_label = ticket_type_label.replace(" ", "_")
        ticket_folder = get_ticket_folder(download_dir, ticket_number)
        new_name = os.path.join(ticket_folder, f"{ticket_number}_{safe_label}.pdf")
        try:
            shutil.move(filepath, new_name)
            return new_name
        except Exception as e:
            log(f"Error moving/renaming file: {e}")
            return None


    def close_export_dialog(driver, ticket):
        def check_still_open(d):
            return deep_find(d, "#download_button")

        still_open = find_across_frames(driver, check_still_open)
        if not still_open:
            driver.switch_to.default_content()
            return

        def find_cancel_or_close(d):
            el = deep_find_by_text(d, "Cancel")
            if el:
                return el
            for sel in ["button[aria-label='Close']", "button[title='Close']", ".close", "[aria-label*='close' i]"]:
                el = deep_find(d, sel)
                if el:
                    return el
            return None

        btn = find_across_frames(driver, find_cancel_or_close)
        if btn:
            try:
                safe_click(driver, btn)
                log(f"  Force-closed 'Export Complete' dialog for {ticket}")
                show_screenshot(driver, f"[{ticket}] Dialog force-closed")
            except Exception as e:
                log(f"  Could not click Cancel/Close on dialog for {ticket}: {e}")
        else:
            log(f"  Dialog still open for {ticket} but no Cancel/Close button found")
        driver.switch_to.default_content()


    def run_automation(sso_id, password, tickets, ticket_type_label, download_dir, headless):
        driver = None
        try:
            try:
                with open(LOG_FILE_PATH, "w", encoding="utf-8") as f:
                    f.write(f"=== Automation run started {now_ist()} ===\n")
            except Exception:
                pass

            driver = build_driver(download_dir, headless=headless)
            login(driver, sso_id, password)

            for ticket in tickets:
                try:
                    log(f"Processing ticket: {ticket}")
                    before_files = _list_files_only(get_candidate_download_dirs(download_dir))

                    download_ticket_pdf(driver, ticket, download_dir)

                    downloaded_file = wait_for_new_download(download_dir, before_files, timeout=120)

                    close_export_dialog(driver, ticket)

                    if downloaded_file:
                        renamed = rename_downloaded_file(downloaded_file, ticket, ticket_type_label, download_dir)
                        if renamed:
                            log(f"Ticket {ticket} downloaded and renamed to: {os.path.relpath(renamed, download_dir)}")
                        else:
                            log(f"Ticket {ticket} downloaded but renaming failed. File was at: {downloaded_file}")
                    else:
                        log(f"Ticket {ticket}: download did not complete within 120s. "
                            f"Checked folders: {get_candidate_download_dirs(download_dir)}")

                except TimeoutException as e:
                    msg = str(e).strip() or "(no additional details)"
                    log(f"Ticket {ticket}: TIMEOUT — {msg}")
                    show_screenshot(driver, f"[{ticket}] TIMEOUT")
                except Exception as ex:
                    log(f"Ticket {ticket}: ERROR — {type(ex).__name__}: {ex}")
                    show_screenshot(driver, f"[{ticket}] ERROR")

            log("All tickets processed.")
            log(f"Full diagnostic log saved to: {LOG_FILE_PATH}")

        except Exception as ex:
            log(f"Error: {str(ex)}")
        finally:
            if driver is not None:
                driver.quit()
            log("Automation completed. Browser closed.")


    # ==========================================================
    # Excel report generation (Change Management report) — extracted from each
    # downloaded ticket PDF, merged into this tool so the report is produced
    # automatically as part of the same run, no separate upload step needed.
    # ==========================================================
    def extract_field_value(text, field_label, length_limit=None):
        try:
            start_index = text.index(field_label) + len(field_label)
            end_index = text.index("\n", start_index)
            value = text[start_index:end_index].strip()
            if length_limit:
                value = value[:length_limit]
            return value
        except ValueError:
            return None


    def extract_requestor_value(text, field_label):
        try:
            start_index = text.index(field_label) + len(field_label)
            end_index = text.index("\n", start_index)
            value = text[start_index:end_index].strip()
            end_paren_index = value.find(")")
            if end_paren_index != -1:
                value = value[: end_paren_index + 1]
            if "Type:" in value:
                value = value.split("Type:")[0].strip()
            return value
        except ValueError:
            return None


    def extract_value_after_line(text, field_label):
        try:
            start_index = text.index(field_label) + len(field_label)
            end_index = text.index("\n", start_index)
            next_line_start = end_index + 1
            next_line_end = text.find("\n", next_line_start)
            if next_line_end == -1:
                next_line_end = len(text)
            value = text[next_line_start:next_line_end].strip()
            return value
        except ValueError:
            return None


    def extract_data_from_pdf(file_bytes: bytes, filename: str):
        text = ""
        approvers = []
        approval_dates = []
        approver_details = []
        change_developer = ""
        change_implementor = ""
        change_implemented_on = ""
        cab_approval_provided = "False"
        cab_approver = None
        cab_approval_date = None

        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for pdf_page in pdf.pages:
                    for table in pdf_page.extract_tables():
                        for row in table:
                            if len(row) >= 10:
                                type_value = (
                                    (" ".join((row[2] or "").split())).lower() if row[2] else None
                                )
                                assigned_to_value = (
                                    " ".join((row[9] or "").split()) if row[9] else None
                                )
                                assigned_end_date = row[8].strip() if row[8] else None

                                if type_value and type_value.startswith("planning"):
                                    change_developer = assigned_to_value
                                elif type_value and type_value.startswith("implement"):
                                    change_implementor = assigned_to_value
                                    change_implemented_on = assigned_end_date
        except Exception as e:
            log(f"    [Excel] pdfplumber failed to extract tables from {filename}: {e}")

        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for pdf_page in pdf.pages:
                    page_text = pdf_page.extract_text()
                    if page_text:
                        text += page_text + "\n"

                    for table in pdf_page.extract_tables():
                        if table and "Approver" in table[0]:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            approved_rows = df[df.iloc[:, 0].str.lower() == "approved"]

                            approvers.extend(approved_rows.iloc[:, 1].tolist())
                            approval_dates.extend(approved_rows.iloc[:, 4].tolist())
                            approver_details.extend(approved_rows.iloc[:, 2].tolist())

                            for idx, detail in enumerate(approved_rows.iloc[:, 2]):
                                normalized_detail = str(detail).replace("\n", "").strip()
                                if "@IT CAB Approvers" in normalized_detail:
                                    cab_approval_provided = "Yes"
                                    cab_approver = approved_rows.iloc[idx, 1]
                                    cab_approval_date = approved_rows.iloc[idx, 4]
                                    break
                            if cab_approver:
                                break
        except Exception as e:
            log(f"    [Excel] pdfplumber failed to extract text from {filename}: {e}")

        ticket_no = extract_field_value(text, "Number:", length_limit=10)
        requestor = extract_requestor_value(text, "Requested by:")
        planned_start_date = extract_field_value(text, "Planned start date:", length_limit=19)
        change_type = extract_field_value(text, "Type:")
        change_description = extract_value_after_line(text, "Short description:")

        change_approved_by = ", ".join(approvers)
        change_approved_on = ", ".join(approval_dates)

        attribute_checks = {
            "Approvals Obtained": "Approval: Approved" in text,
            "Segregation of Duty": change_implementor.strip().lower() != change_developer.strip().lower(),
            "Exception Noted": "",
            "CAB Approval Provided (Yes/No)?": cab_approval_provided,
        }

        return (
            ticket_no,
            requestor,
            planned_start_date,
            change_type,
            change_description,
            change_approved_by,
            change_approved_on,
            change_developer,
            change_implementor,
            change_implemented_on,
            attribute_checks,
            cab_approver,
            cab_approval_date,
            cab_approval_provided,
        )


    def generate_excel_report(data_list) -> bytes:
        df = pd.DataFrame(
            data_list,
            columns=[
                "Sl No", "Change Request Ticket", "Change Description", "Change Requestor",
                "Planned Start Date", "Change Type", "Change Approved By", "Change Approved On",
                "Approvals are obtained for all changes?",
                "Change Implemented By", "Change Implemented On", "Change Developed By",
                "Whether SOD is maintained between the developer and implementor? (Yes/No)",
                "CAB Approval Provided (Yes/No)?",
                "CAB Approver", "CAB Approval Obtained on", "Exception Noted",
            ],
        )

        df["Whether SOD is maintained between the developer and implementor? (Yes/No)"] = df.apply(
            lambda row: "True"
            if str(row["Change Implemented By"]).strip().lower() != str(row["Change Developed By"]).strip().lower()
            else "False",
            axis=1,
        )

        df["Exception Noted"] = df.apply(
            lambda row: "Yes" if row.astype(str).str.contains("False", case=False).any() else "No",
            axis=1,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Testing Table"
        ws.sheet_view.showGridLines = False

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=21)
        control_cell = ws.cell(row=1, column=1, value="Control Description: Testing & approval is required...")
        control_cell.font = Font(name="Source Sans Pro", size=12, bold=True)

        thin_border = Border(
            left=Side(style="thin", color="5B2A86"),
            right=Side(style="thin", color="5B2A86"),
            top=Side(style="thin", color="5B2A86"),
            bottom=Side(style="thin", color="5B2A86"),
        )

        attributes = [
            "Attribute 1", "Approvals are obtained for all changes",
            "Attribute 2", "Segregation of duty is ensured between developers and implementors",
            "Attribute 3", "CAB Approval Provided (Yes/No)?",
        ]
        for i in range(0, len(attributes), 2):
            attr_number = ws.cell(row=3 + i // 2, column=1, value=attributes[i])
            attr_number.font = Font(name="Source Sans Pro", bold=True, color="FFFFFF")
            attr_number.fill = PatternFill(start_color="5B2A86", end_color="5B2A86", fill_type="solid")
            attr_number.border = thin_border
            attr_description = ws.cell(row=3 + i // 2, column=2, value=attributes[i + 1])
            attr_description.border = thin_border
            attr_description.font = Font(name="Source Sans Pro")

        start_row = 8

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                if r_idx > start_row and c_idx == 14:
                    value = ""
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True, color="FFFFFF", name="Source Sans Pro")
                    cell.fill = PatternFill(start_color="5B2A86", end_color="5B2A86", fill_type="solid")
                else:
                    cell.font = Font(name="Source Sans Pro")
                    cell.border = Border(
                        left=Side(style="thin", color="5B2A86"),
                        right=Side(style="thin", color="5B2A86"),
                        top=Side(style="thin", color="5B2A86"),
                        bottom=Side(style="thin", color="5B2A86"),
                    )
                    if c_idx == 14 and r_idx >= start_row + 1:
                        cell.value = f'=IF(O{r_idx}<>"", "Yes", "No")'

        for col in ws.iter_cols(min_row=3, max_row=start_row):
            col_values = [len(str(cell.value)) for cell in col if cell.value is not None]
            if col_values:
                ws.column_dimensions[col[0].column_letter].width = max(col_values) + 2

        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        ws.protection.sheet = True
        ws.protection.enable()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


    def build_excel_report_for_tickets(download_dir: str, tickets, ticket_type_label: str):
        """Reads back each ticket's already-downloaded PDF (from its per-ticket
        folder) and runs it through the same extraction/report logic as the
        standalone Change Management Excel tool -- so the report is produced
        automatically in the SAME run, with no separate upload step. Returns
        the .xlsx bytes, or None if no ticket PDFs could be read."""
        data_list = []
        for i, ticket in enumerate(tickets, start=1):
            ticket_folder = get_ticket_folder(download_dir, ticket)
            pdf_files = sorted(glob.glob(os.path.join(ticket_folder, "*.pdf")))
            if not pdf_files:
                log(f"  [Excel] No PDF found for {ticket}, skipping in Excel report")
                continue
            try:
                with open(pdf_files[0], "rb") as f:
                    file_bytes = f.read()
                (
                    ticket_no, requestor, planned_start_date, change_type,
                    change_description, change_approved_by, change_approved_on,
                    change_developer, change_implementor, change_implemented_on,
                    attribute_checks, cab_approver, cab_approval_date, cab_approval_provided,
                ) = extract_data_from_pdf(file_bytes, os.path.basename(pdf_files[0]))

                # Fall back to the originally-searched ticket number if the PDF
                # text extraction didn't find one (keeps the report row usable
                # even if the ticket's PDF layout differs slightly).
                ticket_no = ticket_no or ticket

                data_list.append(
                    [
                        i, ticket_no, change_description, requestor, planned_start_date,
                        change_type, change_approved_by, change_approved_on,
                        attribute_checks["Approvals Obtained"],
                        change_implementor, change_implemented_on, change_developer,
                        attribute_checks["Segregation of Duty"], cab_approval_provided,
                        cab_approver, cab_approval_date, "",
                    ]
                )
                log(f"  [Excel] Extracted report data for {ticket}")
            except Exception as e:
                log(f"  [Excel] Failed to process {ticket}: {type(e).__name__}: {e}")

        if not data_list:
            log("  [Excel] No valid ticket data to generate report.")
            return None

        return generate_excel_report(data_list)


    def make_zip(download_dir: str, excel_bytes: bytes = None) -> str:
        """Zips the entire per-ticket folder structure, PLUS the Excel
        change-management report at the top level (if available):
        tickets.zip
          change_management_report.xlsx
          CHG0096025/
            CHG0096025_Normal_Ticket.pdf
            attachments/
              CHG0096025_COO File.zip
              ...
          CHG0070418/
            CHG0070418_Normal_Ticket.pdf
            (no attachments/ subfolder if this ticket had none)
        """
        zip_path = os.path.join(download_dir, "tickets.zip")
        with zipfile.ZipFile(zip_path, "w") as zf:
            for ticket_folder in sorted(glob.glob(os.path.join(download_dir, "*"))):
                if not os.path.isdir(ticket_folder):
                    continue
                for root, _dirs, files in os.walk(ticket_folder):
                    for fname in files:
                        if fname == "tickets.zip":
                            continue
                        full_path = os.path.join(root, fname)
                        arcname = os.path.relpath(full_path, download_dir)
                        zf.write(full_path, arcname)
            if excel_bytes:
                zf.writestr("change_management_report.xlsx", excel_bytes)
        return zip_path


    # ==========================================================
    # UI
    # ==========================================================


    step = st.session_state.cm_step

    if step == 0:
        st.header("Confirm Automation Credentials")

        _portal_user = st.session_state.get("current_user", {})
        _portal_sso = _portal_user.get("sso_id", "")
        _portal_pwd = st.session_state.get("plain_password", "")

        if not st.session_state.cm_sso_id:
            st.session_state.cm_sso_id = _portal_sso
        if not st.session_state.cm_password:
            st.session_state.cm_password = _portal_pwd

        st.info(
            "We will reuse the SSO ID and password you used to log into this portal. "
            "You can override them below if you need to run the automation under a "
            "different account."
        )

        st.session_state.cm_sso_id = st.text_input("SSO ID", value=st.session_state.cm_sso_id)
        st.session_state.cm_password = st.text_input("Password", type="password", value=st.session_state.cm_password)

        st.info(
            "Your SSO ID and password will be entered automatically. After that, "
            "you'll just need to approve the push notification on your authenticator "
            "app (PingID) on your mobile device, "
            "no browser interaction needed."
        )

        st.session_state.cm_headless = st.checkbox(
            "Run headless (leave checked for cloud deployment)",
            value=st.session_state.cm_headless,
        )

        if st.button("Next"):
            if st.session_state.cm_sso_id and st.session_state.cm_password:
                st.session_state.cm_step = 1
                st.rerun()
            else:
                st.warning("Please enter your SSO ID and password.")

    elif step == 1:
        labels = TICKET_TYPE_LABELS[st.session_state.cm_platform]
        label_values = list(labels.values())
        default_label = labels[st.session_state.cm_ticket_type_key] if st.session_state.cm_ticket_type_key in labels else label_values[0]
        choice_label = default_label
        st.session_state.cm_ticket_type_key = [k for k, v in labels.items() if v == choice_label][0]

        st.subheader("Paste Ticket Number (in a column, one per line)")
        raw = st.text_area("Tickets", height=200)

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Back"):
                st.session_state.cm_step = 0
                st.rerun()
        with col2:
            if st.button("Load Tickets"):
                tickets = [l.strip() for l in raw.splitlines() if l.strip()]
                if tickets:
                    st.session_state.cm_tickets = tickets
                    st.session_state.cm_step = 2
                    st.rerun()
                else:
                    st.warning("No tickets found in pasted text.")

    elif step == 2:
        st.subheader("Ready to run")

        labels = TICKET_TYPE_LABELS[st.session_state.cm_platform]
        ticket_type_label = labels[st.session_state.cm_ticket_type_key]

        n = len(st.session_state.cm_tickets)
        est = 60 + n * 150
        m, s = divmod(est, 60)
        st.info(f"Estimated time (excluding manual MFA approval): up to {m} min {s} sec")

        st.markdown("**Ticket List**")
        _ticket_rows = "".join(
            f'<div style="padding:6px 12px;border-bottom:1px solid #E5D6F2;'
            f'display:flex;align-items:center;gap:10px;">'
            f'<span style="color:#5B2A86;font-weight:700;font-size:13px;">{i+1}.</span>'
            f'<span style="color:#2B2B2B;font-family:monospace;font-size:14px;">{t}</span>'
            f'</div>'
            for i, t in enumerate(st.session_state.cm_tickets)
        )
        st.markdown(
            "<div style=\"background-color:#FFFFFF;border:1.5px solid #5B2A86;"
            "border-radius:8px;overflow:hidden;margin-bottom:14px;\">"
            + _ticket_rows +
            "</div>",
            unsafe_allow_html=True,
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Back"):
                st.session_state.cm_step = 1
                st.rerun()
        with col2:
            start_clicked = st.button("Start Process")

        col_left, col_right = st.columns([1, 1])
        with col_left:
            st.write("**Live Browser View**")
            _image_placeholder = st.empty()
        with col_right:
            st.write("**Process Log**")
            _log_placeholder = st.empty()
            render_process_log(_log_placeholder, st.session_state.cm_logs)

        if start_clicked:
            tmp_dir = tempfile.mkdtemp()
            st.session_state.cm_download_dir = tmp_dir
            st.session_state.cm_logs = []
            st.session_state.cm_excel_bytes = None
            run_automation(
                st.session_state.cm_sso_id,
                st.session_state.cm_password,
                st.session_state.cm_tickets,
                ticket_type_label,
                tmp_dir,
                st.session_state.cm_headless,
            )
            st.session_state.cm_password = ""

            log("Generating Excel change-management report from downloaded ticket PDFs...")
            with st.spinner("Generating Excel report..."):
                try:
                    st.session_state.cm_excel_bytes = build_excel_report_for_tickets(
                        tmp_dir, st.session_state.cm_tickets, ticket_type_label
                    )
                    if st.session_state.cm_excel_bytes:
                        log("Excel report generated successfully: change_management_report.xlsx")
                    else:
                        log("Excel report could not be generated (no ticket PDFs were readable).")
                except Exception as ex:
                    log(f"Excel report generation failed: {type(ex).__name__}: {ex}")
                    st.session_state.cm_excel_bytes = None

        if os.path.exists(LOG_FILE_PATH):
            with open(LOG_FILE_PATH, "r", encoding="utf-8") as f:
                log_content = f.read()
            st.download_button("Download full diagnostic log (automation_log.txt)",
                                log_content, file_name="automation_log.txt")

        if st.session_state.cm_excel_bytes:
            st.write("### 📊 Change Management Excel Report")
            st.download_button(
                "⬇ Download Excel Report (change_management_report.xlsx)",
                st.session_state.cm_excel_bytes,
                file_name="change_management_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_report",
            )

        if st.session_state.cm_download_dir:
            ticket_folders = sorted([
                d for d in glob.glob(os.path.join(st.session_state.cm_download_dir, "*"))
                if os.path.isdir(d)
            ])

            if ticket_folders:
                st.write("### Downloaded Tickets")
                st.write(f"{len(ticket_folders)} ticket folder(s) ready "
                         f"(each contains the ticket PDF and its attachments, kept separate):")

                for ticket_folder in ticket_folders:
                    ticket_name = os.path.basename(ticket_folder)
                    pdf_files = sorted(glob.glob(os.path.join(ticket_folder, "*.pdf")))
                    attachments_dir = os.path.join(ticket_folder, "attachments")
                    attachment_files = sorted(glob.glob(os.path.join(attachments_dir, "*"))) \
                        if os.path.isdir(attachments_dir) else []

                    with st.expander(
                        f"📁 {ticket_name}  "
                        f"({len(pdf_files)} PDF, {len(attachment_files)} attachment(s))",
                        expanded=False,
                    ):
                        if pdf_files:
                            st.caption("Ticket PDF")
                            for pdf_path in pdf_files:
                                fname = os.path.basename(pdf_path)
                                fsize_kb = os.path.getsize(pdf_path) / 1024
                                with open(pdf_path, "rb") as f:
                                    st.download_button(
                                        f"⬇ {fname} ({fsize_kb:.0f} KB)",
                                        f,
                                        file_name=fname,
                                        mime="application/pdf",
                                        key=f"dl_pdf_{ticket_name}_{fname}",
                                    )
                        else:
                            st.caption("No PDF found for this ticket.")

                        if attachment_files:
                            st.caption("Attachments")
                            for att_path in attachment_files:
                                fname = os.path.basename(att_path)
                                fsize_kb = os.path.getsize(att_path) / 1024
                                with open(att_path, "rb") as f:
                                    st.download_button(
                                        f"⬇ {fname} ({fsize_kb:.0f} KB)",
                                        f,
                                        file_name=fname,
                                        key=f"dl_att_{ticket_name}_{fname}",
                                    )
                        else:
                            st.caption("No attachments for this ticket.")

                zip_path = make_zip(st.session_state.cm_download_dir, st.session_state.cm_excel_bytes)
                zip_label = (
                    "⬇ Download ALL tickets + attachments + Excel report as ZIP "
                    "(organized in per-ticket folders)"
                    if st.session_state.cm_excel_bytes else
                    "⬇ Download ALL tickets + attachments as ZIP "
                    "(organized in per-ticket folders)"
                )
                with open(zip_path, "rb") as f:
                    st.download_button(
                        zip_label,
                        f,
                        file_name="tickets.zip",
                        mime="application/zip",
                        key="dl_zip_all",
                    )
            else:
                st.info("No tickets downloaded yet in this session.")



# ============================================================================
# SECTION 9 — MAIN APP
# ============================================================================
st.set_page_config(page_title="Tech Assisted Audit Tools", page_icon="🔐", layout="wide")
inject_global_css()

if "authenticated" not in st.session_state: st.session_state.authenticated = False
if "current_user" not in st.session_state: st.session_state.current_user = None
if "login_error" not in st.session_state: st.session_state.login_error = ""
if "active_tool" not in st.session_state: st.session_state.active_tool = None
if "portal_view" not in st.session_state: st.session_state.portal_view = "dashboard"


def _logout():
    for key in ["authenticated", "current_user", "login_error", "active_tool", "portal_view", "plain_password", "register_error", "register_success"]:
        if key in st.session_state:
            st.session_state[key] = False if key == "authenticated" else None
    st.session_state.portal_view = "dashboard"
    st.rerun()


def _render_change_password_widget():
    with st.expander("🔑 Change Password"):
        with st.form("change_password_form", clear_on_submit=True):
            current_pwd = st.text_input("Current Password", type="password", key="cp_current")
            new_pwd = st.text_input("New Password", type="password", key="cp_new")
            confirm_pwd = st.text_input("Confirm New Password", type="password", key="cp_confirm")
            cp_submitted = st.form_submit_button("Update Password", use_container_width=True)
        if cp_submitted:
            if new_pwd != confirm_pwd:
                st.error("New password and confirmation do not match.")
            else:
                sso_id = st.session_state.current_user.get("sso_id")
                ok, msg = change_password(sso_id, current_pwd, new_pwd)
                if ok:
                    st.success(msg)
                    st.session_state.plain_password = new_pwd
                else:
                    st.error(msg)


if not st.session_state.authenticated:
    render_login_page()
else:
    _sso_id = st.session_state.current_user.get("sso_id")
    _fresh_user = get_user(_sso_id)
    if _fresh_user:
        _fresh_user["sso_id"] = _sso_id
        st.session_state.current_user = _fresh_user

    user = st.session_state.current_user
    role = user.get("role")
    account_status = user.get("status", STATUS_APPROVED)
    is_admin_role = (role == ROLE_ADMIN)
    pending_count = len(get_pending_users()) if is_admin_role else 0

    with st.sidebar:
        initials = "".join([p[0] for p in user.get("name", "?").split()[:2]]).upper() or "?"
        st.markdown(f"""
            <div class="sidebar-user-card">
                <div class="sidebar-avatar">{initials}</div>
                <div class="sidebar-name">{user.get('name')}</div>
                <div class="sidebar-sso">SSO: {user.get('sso_id')}</div>
                <div class="sidebar-role-pill">{role}</div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown("---")

        _render_change_password_widget()
        st.markdown("---")

        if is_admin_role:
            admin_label = f"Admin Portal ({pending_count} new)" if pending_count else "Admin Portal"
            nav_choice = st.radio("Navigate", options=["My Dashboard", admin_label], index=0 if st.session_state.portal_view == "dashboard" else 1, label_visibility="collapsed")
            st.session_state.portal_view = "dashboard" if nav_choice == "My Dashboard" else "admin"
            st.markdown("---")
        else:
            st.session_state.portal_view = "dashboard"

        if st.button("🚪  Logout", use_container_width=True):
            _logout()

    if account_status == STATUS_PENDING:
        render_pending_screen()
    elif is_admin_role and st.session_state.portal_view == "admin":
        render_admin_portal()
    else:
        render_dashboard()
